from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import calendar

def create_attendance_format(year, month, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月出面表"

    # タイトル
    ws['A1'] = f"{year}年{month}月 出面表"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:AH1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # 月の日数を取得
    _, days_in_month = calendar.monthrange(year, month)

    # ヘッダー行
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 固定列
    ws['A3'] = 'No'
    ws['B3'] = '氏名'
    ws['C3'] = '所属'

    # 日付列
    for day in range(1, days_in_month + 1):
        col = 3 + day  # D列から開始
        cell = ws.cell(3, col, day)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 合計列
    total_col = 3 + days_in_month + 1
    ws.cell(3, total_col, '合計').font = Font(bold=True)
    ws.cell(3, total_col).fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    ws.cell(3, total_col).alignment = Alignment(horizontal='center')

    # ヘッダーのスタイル
    for col in range(1, 4):
        cell = ws.cell(3, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # サンプル作業員データ
    workers = [
        (1, '田中太郎', 'A班'),
        (2, '佐藤花子', 'A班'),
        (3, '鈴木一郎', 'B班'),
        (4, '高橋次郎', 'B班'),
        (5, '渡辺三郎', 'C班'),
    ]

    for row_idx, (no, name, team) in enumerate(workers, 4):
        ws.cell(row_idx, 1, no).border = thin_border
        ws.cell(row_idx, 2, name).border = thin_border
        ws.cell(row_idx, 3, team).border = thin_border

        # 出勤マーク（○）をランダムに配置
        attendance_count = 0
        for day in range(1, days_in_month + 1):
            col = 3 + day
            # 平日は○、土日は空白（簡易版）
            weekday = calendar.weekday(year, month, day)
            if weekday < 5:  # 月〜金
                ws.cell(row_idx, col, '○').alignment = Alignment(horizontal='center')
                attendance_count += 1
            ws.cell(row_idx, col).border = thin_border

        # 合計
        ws.cell(row_idx, total_col, attendance_count).alignment = Alignment(horizontal='center')
        ws.cell(row_idx, total_col).border = thin_border
        ws.cell(row_idx, total_col).font = Font(bold=True)

    # 列幅調整
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    for day in range(1, days_in_month + 1):
        col_letter = ws.cell(3, 3 + day).column_letter
        ws.column_dimensions[col_letter].width = 4

    wb.save(output_path)

if __name__ == "__main__":
    create_attendance_format(
        year=2026,
        month=1,
        output_path="test_attendance.xlsx"
    )
    print("生成完了: test_attendance.xlsx")
