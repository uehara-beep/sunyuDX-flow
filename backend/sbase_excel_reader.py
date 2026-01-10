"""
S-BASE方式のExcel読み取りモジュール
キーワード検索で柔軟にデータを抽出
"""

import openpyxl
import re
from typing import Optional, Dict, List, Any
from pathlib import Path

class SBaseExcelReader:
    """S-BASE方式のExcel読み取りクラス"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
        self.ws = self.wb.active
    
    def find_value_by_keyword(
        self, 
        keyword: str, 
        search_range: int = 30,
        direction: str = "right"
    ) -> Optional[str]:
        """
        キーワードを含むセルを探して、その近くのセルの値を返す
        
        Args:
            keyword: 検索キーワード（例: "工事名", "発注者"）
            search_range: 検索範囲（行数）
            direction: 検索方向 ("right", "below", "both")
        
        Returns:
            見つかった値、または None
        """
        for row in range(1, search_range + 1):
            for col in range(1, 20):  # A列からS列まで
                cell_value = self.ws.cell(row, col).value
                
                if not cell_value:
                    continue
                
                if keyword in str(cell_value):
                    # キーワードが見つかった！
                    
                    # 右隣をチェック
                    if direction in ["right", "both"]:
                        right_cell = self.ws.cell(row, col + 1).value
                        if right_cell and str(right_cell).strip():
                            return str(right_cell).strip()
                        
                        # 右隣が空なら、さらに右を探す
                        for offset in range(2, 10):
                            right_cell = self.ws.cell(row, col + offset).value
                            if right_cell and right_cell != keyword and str(right_cell).strip():
                                return str(right_cell).strip()
                    
                    # 下をチェック
                    if direction in ["below", "both"]:
                        below_cell = self.ws.cell(row + 1, col).value
                        if below_cell and str(below_cell).strip():
                            return str(below_cell).strip()
        
        return None
    
    def find_company_name(self, search_range: int = 20) -> Optional[str]:
        """
        会社名を自動検出
        「株式会社」「NEXCO」などのキーワードで探す
        """
        company_keywords = [
            "株式会社", "有限会社", "合同会社", "合資会社",
            "NEXCO", "JV", "建設", "工業", "組合"
        ]
        
        for row in range(1, search_range + 1):
            for col in range(1, 15):
                cell_value = self.ws.cell(row, col).value
                
                if not cell_value:
                    continue
                
                cell_str = str(cell_value).strip()
                
                # 会社名キーワードを含むか
                for keyword in company_keywords:
                    if keyword in cell_str:
                        # 余計な文字を除去
                        cleaned = cell_str.replace("御中", "").replace("様", "").replace("殿", "").strip()
                        if len(cleaned) > 3:  # 短すぎる文字列は除外
                            return cleaned
        
        return None
    
    def extract_amount(self, keyword: str = "合計") -> Optional[int]:
        """
        金額を抽出
        数字だけを取り出して整数に変換
        """
        amount_str = self.find_value_by_keyword(keyword, direction="both")
        
        if not amount_str:
            # 他のキーワードでも試す
            for kw in ["小計", "総額", "工事金額", "請負金額"]:
                amount_str = self.find_value_by_keyword(kw, direction="both")
                if amount_str:
                    break
        
        if amount_str:
            try:
                # 数字だけ抽出
                numbers = re.sub(r'[^\d]', '', str(amount_str))
                if numbers:
                    return int(numbers)
            except:
                pass
        
        return None
    
    def extract_all_sheets_items(self) -> List[Dict[str, Any]]:
        """
        全シートから内訳明細を抽出
        S-BASE方式：キーワードで自動検出
        """
        items = []
        item_keywords = [
            "工", "費", "作業", "施工", "設置", "撤去", "補修",
            "防水", "舗装", "切断", "清掃", "運搬", "処分"
        ]
        
        for sheet_name in self.wb.sheetnames:
            # 「条件書」などのシートはスキップ
            if any(skip in sheet_name for skip in ["条件", "表紙", "備考"]):
                continue
            
            ws = self.wb[sheet_name]
            
            for row in range(1, 200):  # 最大200行
                for col in range(1, 15):  # A-N列
                    cell_value = ws.cell(row, col).value
                    
                    if not cell_value:
                        continue
                    
                    cell_str = str(cell_value).strip()
                    
                    # 項目名っぽいか？
                    if any(kw in cell_str for kw in item_keywords):
                        # 同じ行で金額を探す
                        amount = None
                        for amt_col in range(col, min(col + 10, 20)):
                            amt_value = ws.cell(row, amt_col).value
                            if isinstance(amt_value, (int, float)) and amt_value > 1000:
                                amount = float(amt_value)
                                break
                        
                        # 数量も探す
                        quantity = None
                        unit = None
                        for qty_col in range(col, min(col + 10, 20)):
                            qty_value = ws.cell(row, qty_col).value
                            if isinstance(qty_value, (int, float)) and 0 < qty_value < 10000:
                                quantity = float(qty_value)
                                # 単位も探す
                                unit_value = ws.cell(row, qty_col + 1).value
                                if unit_value and str(unit_value) in ["m", "m2", "m3", "式", "箇所", "本", "枚"]:
                                    unit = str(unit_value)
                                break
                        
                        if amount and amount > 1000:
                            items.append({
                                "id": f"{sheet_name}_{row}",
                                "sheet_name": sheet_name,
                                "item_name": cell_str,
                                "quantity": quantity if quantity else 1.0,
                                "unit": unit if unit else "式",
                                "amount": amount
                            })
        
        # 重複除去（同じ項目名+金額）
        seen = set()
        unique_items = []
        for item in items:
            key = (item["item_name"], item["amount"])
            if key not in seen:
                seen.add(key)
                unique_items.append(item)
        
        return unique_items[:20]  # 最大20件
    
    def extract_project_info(self) -> Dict[str, Any]:
        """
        プロジェクト情報を完全抽出
        S-BASE方式：キーワード検索で柔軟に対応
        """
        # 工事名を探す（複数パターン）
        project_name = None
        for keyword in ["工事名", "工 事 名", "件名", "工事件名", "現場名"]:
            project_name = self.find_value_by_keyword(keyword, direction="both")
            if project_name:
                break
        
        # 発注者を探す
        client_name = None
        for keyword in ["発注者", "注文者", "御中", "様", "殿"]:
            client_name = self.find_value_by_keyword(keyword, direction="both")
            if client_name:
                # 余計な文字を除去
                client_name = client_name.replace("御中", "").replace("様", "").replace("殿", "").strip()
                break
        
        # 会社名が見つからなければ自動検出
        if not client_name:
            client_name = self.find_company_name()
        
        # 工事場所を探す
        location = None
        for keyword in ["工事場所", "場所", "施工場所", "工事箇所", "所在地"]:
            location = self.find_value_by_keyword(keyword, direction="both")
            if location:
                break
        
        # 工期を探す
        period = None
        for keyword in ["工期", "工事期間", "契約工期"]:
            period = self.find_value_by_keyword(keyword, direction="both")
            if period:
                break
        
        # 金額を探す
        amount = self.extract_amount()
        
        # デフォルト値
        if not project_name:
            project_name = "工事名（自動取得失敗）"
        if not client_name:
            client_name = "発注者（自動取得失敗）"
        if not location:
            location = "場所（自動取得失敗）"
        if not amount:
            amount = 10000000
        
        return {
            "project_name": project_name,
            "client_name": client_name,
            "location": location,
            "period": period,
            "amount": amount
        }

def extract_from_excel(file_path: str) -> Dict[str, Any]:
    """
    Excel見積書から情報を抽出（S-BASE方式）
    
    Args:
        file_path: Excelファイルのパス
    
    Returns:
        抽出された情報の辞書
    """
    reader = SBaseExcelReader(file_path)
    
    # プロジェクト基本情報
    project_info = reader.extract_project_info()
    
    # 内訳明細
    items = reader.extract_all_sheets_items()
    
    return {
        **project_info,
        "items": items,
        "sheets": reader.wb.sheetnames
    }

# テスト用
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        result = extract_from_excel(sys.argv[1])
        print("=" * 60)
        print("📊 S-BASE方式のExcel解析結果")
        print("=" * 60)
        print(f"工事名: {result['project_name']}")
        print(f"発注者: {result['client_name']}")
        print(f"場所: {result['location']}")
        print(f"工期: {result.get('period', 'なし')}")
        print(f"金額: ¥{result['amount']:,}")
        print(f"\nシート数: {len(result['sheets'])}")
        print(f"内訳明細数: {len(result['items'])}")
        print("\n📋 内訳明細（最初の5件）:")
        for i, item in enumerate(result['items'][:5], 1):
            print(f"  {i}. [{item['sheet_name']}] {item['item_name']} - ¥{item['amount']:,}")
