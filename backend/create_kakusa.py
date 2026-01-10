from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import date

def create_kakusa_format(estimate_number, project_name, client_name, location, subtotal, output_path):
    wb = Workbook()

    # 表紙
    ws = wb.active
    ws.title = "表紙"
    ws['A1'] = "工事見積書"
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')

    ws['A3'] = "見積番号"
    ws['B3'] = estimate_number
    ws['A4'] = "発注者"
    ws['B4'] = client_name
    ws['A5'] = "工事名"
    ws['B5'] = project_name
    ws['A6'] = "工事場所"
    ws['B6'] = location
    ws['A8'] = "見積金額（税込）"
    ws['B8'] = f"¥{subtotal:,.0f}"
    ws['B8'].font = Font(size=14, bold=True)
    ws['A10'] = "見積日"
    ws['B10'] = date.today().strftime('%Y年%m月%d日')
    ws['A12'] = "株式会社サンユウテック"
    ws['A12'].font = Font(size=12, bold=True)

    # 内訳明細
    ws2 = wb.create_sheet("内訳明細")
    headers = ['No', '項目', '数量', '単位', '単価', '金額', '備考']
    for col, h in enumerate(headers, 1):
        c = ws2.cell(1, col, h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

    items = [
        (1, "準備工", 1, "式", 50000, 50000),
        (2, "WJ工事", 100, "m2", 5000, 500000),
        (3, "舗装工事", 100, "m2", 3000, 300000),
        (4, "片付け工", 1, "式", 30000, 30000),
    ]

    for row, item in enumerate(items, 2):
        for col, val in enumerate(item, 1):
            c = ws2.cell(row, col, val)
            if col in [3, 5, 6]:
                c.alignment = Alignment(horizontal='right')

    ws2.cell(len(items)+3, 1, "合計").font = Font(bold=True)
    ws2.cell(len(items)+3, 6, subtotal).font = Font(bold=True)

    # KAKUSA
    ws3 = wb.create_sheet("KAKUSA")
    ws3['A1'] = "KAKUSA形式見積書"
    ws3['A1'].font = Font(size=16, bold=True)
    ws3['A3'] = "工事名"
    ws3['B3'] = project_name
    ws3['A4'] = "見積番号"
    ws3['B4'] = estimate_number
    ws3['A5'] = "合計金額"
    ws3['B5'] = f"¥{subtotal:,.0f}"

    wb.save(output_path)
