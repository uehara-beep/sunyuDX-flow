from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import date

def create_ledger_format(project_name, output_path):
    wb = Workbook()

    # 5科目のシートを作成
    categories = [
        ("材料費", "FFA500"),
        ("労務費", "4169E1"),
        ("機械費", "32CD32"),
        ("外注費", "DC143C"),
        ("経費", "9370DB")
    ]

    # 最初のシートを削除
    wb.remove(wb.active)

    for category_name, color in categories:
        ws = wb.create_sheet(category_name)

        # ヘッダー
        ws['A1'] = f"工事台帳 - {category_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')

        ws['A2'] = f"工事名: {project_name}"
        ws['A2'].font = Font(size=12)
        ws.merge_cells('A2:H2')

        ws['A3'] = f"作成日: {date.today().strftime('%Y年%m月%d日')}"
        ws.merge_cells('A3:H3')

        # テーブルヘッダー
        headers = ['日付', '項目', '数量', '単位', '単価', '金額', '備考', '承認']
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(5, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # サンプルデータ
        sample_data = [
            (date.today().strftime('%Y/%m/%d'), 'サンプル項目1', 10, '個', 1000, 10000, '', ''),
            (date.today().strftime('%Y/%m/%d'), 'サンプル項目2', 5, 'm', 2000, 10000, '', ''),
            (date.today().strftime('%Y/%m/%d'), 'サンプル項目3', 1, '式', 50000, 50000, '', ''),
        ]

        for row_idx, data in enumerate(sample_data, 6):
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row_idx, col_idx, value)
                cell.border = thin_border
                if col_idx in [3, 5, 6]:  # 数量、単価、金額
                    cell.alignment = Alignment(horizontal='right')

        # 合計行
        total_row = 6 + len(sample_data) + 1
        ws.cell(total_row, 1, '合計').font = Font(bold=True)
        ws.cell(total_row, 6, 70000).font = Font(bold=True)
        ws.cell(total_row, 6).alignment = Alignment(horizontal='right')

        # 列幅調整
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 10

    wb.save(output_path)

if __name__ == "__main__":
    create_ledger_format(
        project_name="テスト工事",
        output_path="test_ledger.xlsx"
    )
    print("生成完了: test_ledger.xlsx")
