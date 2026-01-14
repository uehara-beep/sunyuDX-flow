"""
sunyuDX-flow Backend API
S-BASE方式の完全実装版
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session
import pandas as pd
import openpyxl
from datetime import datetime
import uuid
import os
from pathlib import Path
import json
import re
import requests
from dotenv import load_dotenv

# 環境変数読み込み
load_dotenv()

# Database imports
from database import get_db, engine
from models import (
    Project as ProjectModel,
    Estimate as EstimateModel,
    CostRecord as CostRecordModel,
    DailyReport as DailyReportModel,
    DailyReportItem as DailyReportItemModel,
    Invoice as InvoiceModel,
    EstimateImport as EstimateImportModel,
    EstimateLine as EstimateLineModel,
    Attachment as AttachmentModel,
    Base
)
import hashlib

# FastAPIアプリケーション
app = FastAPI(
    title="sunyuDX-flow API",
    description="S-BASE方式の次世代建設DXプラットフォーム",
    version="1.0.0"
)

# CORS設定
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001", "http://localhost:3002", "http://localhost:3003", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =====================================
# Health Check
# =====================================
@app.get("/health")
async def health_check():
    """ヘルスチェックエンドポイント"""
    return {"status": "ok", "service": "sunyuDX-flow API", "version": "1.0.0"}

@app.get("/health/db")
async def health_check_db():
    """データベースヘルスチェック"""
    try:
        from database import engine
        from sqlalchemy import text
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return {"status": "ok", "database": "connected"}
    except Exception as e:
        return JSONResponse(
            status_code=503,
            content={"status": "error", "database": str(e)}
        )

# ディレクトリ設定
UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# =====================================
# データモデル (Pydantic)
# =====================================

class CostItem(BaseModel):
    """原価項目"""
    id: str
    name: str
    quantity: float
    unit: str
    unit_price: float
    amount: float

class BudgetCategory(BaseModel):
    """予算カテゴリ（5科目）"""
    name: str
    items: List[CostItem]
    total: float

class Budget(BaseModel):
    """実行予算"""
    id: str
    project_name: str
    client_name: str
    material: BudgetCategory
    labor: BudgetCategory
    equipment: BudgetCategory
    subcontract: BudgetCategory
    expense: BudgetCategory
    budget_total: float
    profit_rate: float
    profit_amount: float
    estimate_amount: float
    created_at: datetime

class EstimateBreakdown(BaseModel):
    """見積内訳"""
    id: str
    sheet_name: str
    item_name: str
    amount: float

class Project(BaseModel):
    """工事プロジェクト"""
    id: str
    name: str
    client: str
    contract_amount: float
    budget_amount: float
    actual_cost: float
    profit_rate: float
    progress: float
    status: str  # active, warning, danger
    created_at: datetime

class ProjectCreate(BaseModel):
    """工事プロジェクト作成用"""
    name: str
    client_name: str
    contract_amount: float = 0
    budget_amount: float = 0
    status: str = "active"

class EstimateCreate(BaseModel):
    """見積作成用"""
    name: str
    amount: float = 0

class CostCreate(BaseModel):
    """原価作成用"""
    category: str  # labor, subcontract, material, machine, expense
    amount: float
    note: Optional[str] = None


class DailyReportItemCreate(BaseModel):
    """日報明細作成用"""
    worker_name: str
    hours: float = 8.0
    wage_rate: int = 15000


class DailyReportCreate(BaseModel):
    """日報作成用"""
    work_date: str  # YYYY-MM-DD
    foreman_name: str
    notes: Optional[str] = None
    items: List[DailyReportItemCreate]


class InvoiceCreate(BaseModel):
    """請求作成用"""
    billing_month: str  # YYYY-MM
    amount: float
    notes: Optional[str] = None


class InvoiceStatusUpdate(BaseModel):
    """請求ステータス更新用"""
    status: str  # draft, issued, paid


class SimpleBudget(BaseModel):
    """予算一覧用シンプルモデル"""
    id: Optional[str] = None
    projectName: str
    projectCode: str
    customerName: str
    totalAmount: float
    laborCost: float
    outsourceCost: float
    materialCost: float
    equipmentCost: float
    expenseCost: float
    profitRate: float
    status: str = "draft"  # draft, pending, approved, rejected
    createdBy: str
    createdAt: Optional[str] = None
    updatedAt: Optional[str] = None

# =====================================
# インメモリデータストア（デモ用）
# =====================================

budgets_db: Dict[str, Budget] = {}
projects_db: Dict[str, Project] = {}
simple_budgets_db: Dict[str, dict] = {}

# =====================================
# ユーティリティ関数
# =====================================

def analyze_excel_file(file_path: Path) -> List[EstimateBreakdown]:
    """
    Excelファイルを解析して内訳明細を抽出
    S-BASE方式: 全シートをスキャンして自動抽出
    """
    breakdowns = []
    
    try:
        # openpyxlでExcelを開く
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 全シートをスキャン
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # キーワード検索: 「金額」「合計」「小計」を含む行を探す
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100), start=1):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # 「○○工」や「○○費」などを検索
                        if any(keyword in cell.value for keyword in ['工', '費', '作業']):
                            # 同じ行の金額を探す
                            amount = None
                            for amt_cell in row:
                                if isinstance(amt_cell.value, (int, float)) and amt_cell.value > 1000:
                                    amount = float(amt_cell.value)
                                    break
                            
                            if amount:
                                breakdown = EstimateBreakdown(
                                    id=str(uuid.uuid4()),
                                    sheet_name=sheet_name,
                                    item_name=str(cell.value),
                                    amount=amount
                                )
                                breakdowns.append(breakdown)
                                break  # 同じ行で複数ヒットしないように
        
        # 重複を除去
        seen = set()
        unique_breakdowns = []
        for b in breakdowns:
            key = (b.sheet_name, b.item_name, b.amount)
            if key not in seen:
                seen.add(key)
                unique_breakdowns.append(b)
        
        return unique_breakdowns[:10]  # 最大10件
        
    except Exception as e:
        print(f"Excel解析エラー: {e}")
        return []

def classify_cost_item(item_name: str, amount: float) -> str:
    """
    AIによる原価分類（簡易版）
    本番環境ではClaude APIを使用
    """
    item_lower = item_name.lower()
    
    # キーワードベースの分類
    if any(word in item_lower for word in ['材料', '資材', '管', 'パイプ', 'コンクリート']):
        return 'material'
    elif any(word in item_lower for word in ['労務', '人工', '作業員', '技術者']):
        return 'labor'
    elif any(word in item_lower for word in ['機械', '重機', 'リース', 'クレーン']):
        return 'equipment'
    elif any(word in item_lower for word in ['外注', '下請', '業者']):
        return 'subcontract'
    else:
        return 'expense'

def generate_estimate_pdf(budget: Budget) -> Path:
    """
    見積書PDFを生成
    S-BASE方式: ReportLabを使用した高品質PDF
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    # 日本語フォント設定（システムにある場合）
    try:
        pdfmetrics.registerFont(TTFont('Japanese', '/usr/share/fonts/truetype/fonts-japanese-gothic.ttf'))
    except:
        pass  # フォントがない場合はデフォルト
    
    # PDF出力パス
    pdf_path = OUTPUT_DIR / f"estimate_{budget.id}.pdf"
    
    # PDFドキュメント作成
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # タイトル
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0066cc'),
        spaceAfter=20,
    )
    story.append(Paragraph('見積書', title_style))
    story.append(Spacer(1, 10*mm))
    
    # プロジェクト情報
    info_data = [
        ['工事名', budget.project_name],
        ['発注者', budget.client_name],
        ['作成日', budget.created_at.strftime('%Y年%m月%d日')],
    ]
    info_table = Table(info_data, colWidths=[50*mm, 100*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 10*mm))
    
    # 金額サマリー
    summary_data = [
        ['実行予算合計', f'¥{budget.budget_total:,.0f}'],
        [f'粗利率 {budget.profit_rate}%', f'¥{budget.profit_amount:,.0f}'],
        ['見積金額', f'¥{budget.estimate_amount:,.0f}'],
    ]
    summary_table = Table(summary_data, colWidths=[80*mm, 70*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#0066cc')),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.white),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 1), 14),
        ('FONTSIZE', (0, 2), (-1, 2), 18),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 10*mm))
    
    # 5科目内訳
    story.append(Paragraph('内訳明細', styles['Heading2']))
    story.append(Spacer(1, 5*mm))
    
    for category_name, category in [
        ('材料費', budget.material),
        ('労務費', budget.labor),
        ('機械費', budget.equipment),
        ('外注費', budget.subcontract),
        ('経費', budget.expense),
    ]:
        if category.total > 0:
            story.append(Paragraph(f'{category_name}: ¥{category.total:,.0f}', styles['Heading3']))
            
            if category.items:
                item_data = [['項目', '数量', '単位', '単価', '金額']]
                for item in category.items[:5]:  # 最大5件
                    item_data.append([
                        item.name,
                        f'{item.quantity:.1f}',
                        item.unit,
                        f'¥{item.unit_price:,.0f}',
                        f'¥{item.amount:,.0f}',
                    ])
                
                item_table = Table(item_data, colWidths=[60*mm, 20*mm, 20*mm, 30*mm, 30*mm])
                item_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                story.append(item_table)
            
            story.append(Spacer(1, 5*mm))
    
    # PDF生成
    doc.build(story)
    
    return pdf_path

# =====================================
# APIエンドポイント
# =====================================

@app.get("/")
async def root():
    """ヘルスチェック"""
    return {
        "status": "ok",
        "message": "sunyuDX-flow API is running",
        "version": "1.0.0"
    }

@app.post("/api/estimate/upload")
async def upload_estimate(file: UploadFile = File(...)):
    """
    見積書Excelをアップロードして解析
    """
    try:
        # ファイル保存
        file_path = UPLOAD_DIR / f"{uuid.uuid4()}_{file.filename}"
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Excel解析
        breakdowns = analyze_excel_file(file_path)
        
        # 一時ファイル削除
        file_path.unlink()
        
        return {
            "status": "success",
            "breakdowns": [b.dict() for b in breakdowns]
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析エラー: {str(e)}")


# =====================================
# Excel取込・コミットAPI
# =====================================

# ヘッダー列の同義語定義（正規化後のキーワードで contains マッチ）
# より多くのパターンに対応するため、一般的な日本語の見積・予算書の表現を網羅
# 注意: 単一文字のシノニムは誤マッチが多いので避ける
HEADER_SYNONYMS = {
    'name': ['名称', '項目名', '工種', '品名', 'name', '名前', '内容', '工事名', '作業内容',
             '費目', '摘要', '種別', '科目', '細目', '中項目', '大項目', '工事内容',
             '種目', '区分', '品目', '件名', '細別', '項目'],
    'breakdown': ['内訳', '仕様', '規格', 'breakdown', 'spec', '明細', '詳細', '形状寸法',
                  '形状', '寸法', '規格寸法', '仕様規格'],
    'qty': ['数量', '出来高', 'qty', 'quantity', '員数', '個数', '本数', '台数',
            '式数', '人数', '回数', '日数', '枚数'],
    'unit': ['単位', 'unit'],
    'unit_price': ['単価', '予算単価', 'unit_price', 'price', '単金', '定価'],
    'amount': ['金額', '請求額', '価格', 'amount', 'total'],
    'note': ['備考', 'note', 'メモ', '注記', 'remark', 'remarks', '特記', '附記'],
}


def normalize_header_text(text) -> str:
    """
    ヘッダーセル文字列を正規化（contains判定用）
    - strip / 全角→半角 / 記号除去 / 空白除去
    例: "数量(㎥)" → "数量"
    例: "単価（円）" → "単価"
    例: "金額（税抜）" → "金額"
    """
    if text is None:
        return ''
    s = str(text).strip()
    if not s:
        return ''

    # 全角英数→半角
    trans_table = str.maketrans(
        'ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ０１２３４５６７８９',
        'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789'
    )
    s = s.translate(trans_table)

    # 記号を除去（より多くの記号に対応）
    remove_chars = '()（）【】[]「」『』〔〕<>＜＞{}｛｝：:・_-－ー／/\\＼、。，,.;；!！?？#＃*＊@＠~～^＾+＋=＝|｜'
    for c in remove_chars:
        s = s.replace(c, '')

    # 空白除去（全角・半角・タブ・改行）
    s = re.sub(r'\s+', '', s)

    # 単位記号・通貨記号を除去
    unit_patterns = ['㎥', '㎡', 'm3', 'm2', 'M3', 'M2', '円', '¥', '￥',
                     '税込', '税抜', '込', '抜', 'kg', 'KG', 't', 'T',
                     '本', '個', '台', '式', '人', '日', '回', '枚', '件',
                     'm', 'M', 'ℓ', 'L', '号', '番']
    for p in unit_patterns:
        s = s.replace(p, '')

    return s.lower()


def normalize_number(value) -> Optional[float]:
    """
    数値を正規化（カンマ除去、全角→半角、通貨記号除去、"1式"対応など）
    対応パターン: "42㎡", "42.0 m2", "42　", "４２", "1,234", "一式" など
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    # 全角数字・記号→半角
    trans_table = str.maketrans('０１２３４５６７８９．，－＋', '0123456789.,-+')
    s = s.translate(trans_table)

    # 空白除去（全角・半角）- 最初に実行
    s = re.sub(r'\s+', '', s)

    # 通貨記号除去
    for currency in ['¥', '￥', '$', '＄', '円', '€', '\\']:
        s = s.replace(currency, '')

    # カンマ除去
    s = s.replace(',', '').replace('、', '')

    # 括弧で囲まれた負数を処理 (123) -> -123
    s = s.strip()
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    # 全角括弧も対応
    if s.startswith('（') and s.endswith('）'):
        s = '-' + s[1:-1]

    # "1式" "一式" などから数値抽出
    if '式' in s:
        s = s.replace('式', '').strip()
        if not s or s in ['一', '１', '1', '']:
            return 1.0

    # 漢数字の一を1に
    s = s.replace('一', '1')

    # 単位文字を除去（より網羅的に）
    unit_patterns = [
        '㎡', '㎥', 'm2', 'm3', 'M2', 'M3',  # 面積・体積
        'km', 'cm', 'mm', 'm', 'ｍ', 'M',     # 長さ
        'kg', 'KG', 'Kg', 't', 'T', 'g',       # 重さ
        'ℓ', 'L', 'l', 'ml', 'ML',             # 容量
        '個', '本', '台', '件', '人', '回', '日', '時間', '枚', '組', '箇所', '基', '棟', '戸',  # 数量単位
        'セット', 'SET', 'set', 'ケース', 'ロット',
    ]
    for unit in unit_patterns:
        s = s.replace(unit, '')

    # 空白除去（最終）
    s = s.strip()

    # 空になった場合
    if not s:
        return None

    # 数値変換を試みる
    try:
        return float(s)
    except ValueError:
        return None


def find_header_column(header_text: str, field: str) -> bool:
    """
    ヘッダーテキストが指定フィールドの同義語に該当するか判定（contains方式）
    正規化後のテキストで部分一致を行う
    """
    normalized = normalize_header_text(header_text)
    if not normalized:
        return False

    for synonym in HEADER_SYNONYMS.get(field, []):
        syn_normalized = normalize_header_text(synonym)
        # 双方向でcontains判定
        if syn_normalized in normalized or normalized in syn_normalized:
            return True
    return False


def parse_excel_to_lines(file_path: Path) -> dict:
    """
    Excelファイルを解析して統一明細形式で抽出
    診断情報付きで返す
    """
    result = {
        'lines': [],
        'sheets_processed': [],
        'sheets_skipped': [],
        'detected_headers': {},
        'missing_columns': [],
        'parse_errors': [],
        'value_stats': {},  # 欠損率統計
        'reason': None,  # 0件時の理由コード
    }

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_result = parse_sheet(sheet, sheet_name)

            if sheet_result['lines']:
                result['lines'].extend(sheet_result['lines'])
                result['sheets_processed'].append({
                    'name': sheet_name,
                    'line_count': len(sheet_result['lines']),
                    'header_row': sheet_result['header_row'],
                    'detected_columns': list(sheet_result['header_map'].keys()),
                })
                # 検出されたヘッダーを記録
                if not result['detected_headers']:
                    result['detected_headers'] = sheet_result['header_map']
            else:
                result['sheets_skipped'].append({
                    'name': sheet_name,
                    'reason': sheet_result.get('skip_reason', 'ヘッダー未検出'),
                    'candidates': sheet_result.get('all_candidates', [])[:3],  # 上位3候補
                })

        wb.close()

        # 必須カラムチェック
        if result['lines']:
            required = ['name', 'amount']
            for col in required:
                if col not in result['detected_headers']:
                    result['missing_columns'].append(col)

            # 欠損率統計を計算
            total = len(result['lines'])
            stats = {
                'qty_missing': sum(1 for l in result['lines'] if l.get('qty') is None),
                'unit_missing': sum(1 for l in result['lines'] if not l.get('unit')),
                'unit_price_missing': sum(1 for l in result['lines'] if l.get('unit_price') is None),
                'amount_missing': sum(1 for l in result['lines'] if l.get('amount') is None),
            }
            result['value_stats'] = {
                'total_lines': total,
                'qty_missing_rate': f"{stats['qty_missing'] / total * 100:.1f}%" if total > 0 else "0%",
                'unit_missing_rate': f"{stats['unit_missing'] / total * 100:.1f}%" if total > 0 else "0%",
                'unit_price_missing_rate': f"{stats['unit_price_missing'] / total * 100:.1f}%" if total > 0 else "0%",
                'amount_missing_rate': f"{stats['amount_missing'] / total * 100:.1f}%" if total > 0 else "0%",
            }

            # ログ出力
            print(f"[Excel解析] {file_path.name}: {total}行取込")
            print(f"  数量欠損: {result['value_stats']['qty_missing_rate']}")
            print(f"  単位欠損: {result['value_stats']['unit_missing_rate']}")
            print(f"  単価欠損: {result['value_stats']['unit_price_missing_rate']}")
            print(f"  金額欠損: {result['value_stats']['amount_missing_rate']}")

        else:
            # 0件の理由を判定
            if result['parse_errors']:
                result['reason'] = 'parse_error'
            elif not result['sheets_processed'] and not result['sheets_skipped']:
                result['reason'] = 'empty_workbook'
            elif all(s.get('reason', '').startswith('ヘッダー行が見つかりません') for s in result['sheets_skipped']):
                result['reason'] = 'header_not_found'
            elif any('必須カラム' in s.get('reason', '') for s in result['sheets_skipped']):
                result['reason'] = 'required_columns_missing'
            else:
                result['reason'] = 'no_data_rows'

    except Exception as e:
        result['parse_errors'].append(f"Excel解析エラー: {str(e)}")
        result['reason'] = 'parse_error'
        import traceback
        print(f"Excel解析エラー: {traceback.format_exc()}")

    return result


def get_merged_cell_value(sheet, row_idx: int, col_idx: int) -> str:
    """
    結合セルの場合、左上セルの値を返す
    """
    cell = sheet.cell(row=row_idx, column=col_idx + 1)  # 1-indexed
    if cell.value is not None:
        return str(cell.value).strip()

    # 結合セルかチェック
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row_idx <= merged_range.max_row and
            merged_range.min_col <= col_idx + 1 <= merged_range.max_col):
            # 左上セルの値を取得
            top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return str(top_left.value or '').strip()

    return ''


def get_row_values_with_merge(sheet, row_idx: int, max_col: int = 20) -> list:
    """
    行の値を取得（結合セルを考慮）
    """
    values = []
    for col_idx in range(max_col):
        val = get_merged_cell_value(sheet, row_idx, col_idx)
        values.append(val)
    return values


def try_map_header_row(values: list) -> dict:
    """
    行の値からヘッダーマッピングを試行
    """
    header_map = {}
    for col_idx, val in enumerate(values):
        if not val:
            continue
        for field in HEADER_SYNONYMS.keys():
            if find_header_column(val, field):
                if field not in header_map:
                    header_map[field] = col_idx
                break
    return header_map


def parse_sheet(sheet, sheet_name: str) -> dict:
    """
    シートを解析して明細行を抽出
    - 結合セル対応
    - 2段ヘッダー対応（r行とr+1行を合成）
    - contains方式のマッチング
    - キー列が2つ以上でヘッダー採用（名称なしでもOK）
    - 複数ヘッダーセクション対応（1シート内に複数の表がある場合）
    """
    print(f"\n[parse_sheet] シート解析開始: {sheet_name}")

    result = {
        'lines': [],
        'header_row': None,
        'header_rows': [],  # 2段ヘッダーの場合 [r, r+1]
        'header_map': {},
        'skip_reason': None,
        'all_candidates': [],  # 検出候補の診断用
        'sections': [],  # 検出した全セクション情報
    }

    # キーカラム（これらが2つ以上あればヘッダー行と判定）
    # nameなしでも qty+unit_price や qty+amount でOK
    key_columns = ['name', 'qty', 'unit_price', 'amount']

    # ヘッダー行を探す（最大100行まで - 複数セクション対応のため拡大）
    max_header_search = 100
    max_col = min(sheet.max_column or 20, 30)

    print(f"  max_col={max_col}, max_header_search={max_header_search}")

    # 全行のキャッシュ（結合セル対応）
    row_cache = {}
    for row_idx in range(1, max_header_search + 2):  # +1 for 2-row header
        try:
            row_cache[row_idx] = get_row_values_with_merge(sheet, row_idx, max_col)
        except Exception as e:
            print(f"  [警告] 行{row_idx}の読取エラー: {e}")
            row_cache[row_idx] = [''] * max_col

    # ヘッダー候補を全て収集
    header_candidates = []

    for row_idx in range(1, max_header_search + 1):
        values = row_cache.get(row_idx, [])

        # 空行はスキップ
        if not any(values):
            continue

        # 単一行でヘッダーマッピング試行
        header_map = try_map_header_row(values)
        key_count = sum(1 for k in key_columns if k in header_map)

        # 候補として記録（診断用に先頭10セル値も保存）
        raw_cells = [v[:15] for v in values[:10] if v]  # 先頭10セル、各15文字まで
        result['all_candidates'].append({
            'row': row_idx,
            'columns': list(header_map.keys()),
            'key_count': key_count,
            'raw_cells': raw_cells[:5],  # 診断用に5セルまで
        })

        # デバッグ: キー列が1つ以上見つかった行を表示
        if key_count >= 1:
            print(f"  行{row_idx}: キー{key_count}個 {list(header_map.keys())} raw={raw_cells[:3]}")

        # キーカラムが2つ以上ある行をヘッダー候補として記録
        if key_count >= 2:
            header_candidates.append({
                'row': row_idx,
                'rows': [row_idx],
                'map': header_map,
                'key_count': key_count,
                'total_count': len(header_map),
                'is_two_row': False
            })

        # 2段ヘッダー試行: 単一行が弱い（キー列1つのみ）場合、次の行と合成
        elif key_count == 1 and row_idx < max_header_search:
            next_values = row_cache.get(row_idx + 1, [])
            next_map = try_map_header_row(next_values)
            next_key_count = sum(1 for k in key_columns if k in next_map)

            # 次の行のキー列が多い場合、次の行をベースにして現在の行を補完
            if next_key_count >= key_count:
                combined_map = dict(next_map)  # 次の行をベース
                for field, col_idx in header_map.items():
                    if field not in combined_map:
                        combined_map[field] = col_idx
            else:
                combined_map = dict(header_map)  # 現在の行をベース
                for field, col_idx in next_map.items():
                    if field not in combined_map:
                        combined_map[field] = col_idx

            combined_key_count = sum(1 for k in key_columns if k in combined_map)

            if combined_key_count >= 2:
                header_candidates.append({
                    'row': row_idx,
                    'rows': [row_idx, row_idx + 1],
                    'map': combined_map,
                    'key_count': combined_key_count,
                    'total_count': len(combined_map),
                    'is_two_row': True
                })

    # ヘッダー候補がない場合
    if not header_candidates:
        top_candidates = sorted(result['all_candidates'], key=lambda x: x['key_count'], reverse=True)[:5]
        diag_lines = []
        for c in top_candidates:
            raw_str = ', '.join(c.get('raw_cells', [])[:3])
            diag_lines.append(f"行{c['row']}: キー{c['key_count']}個 {c['columns']} [{raw_str}]")
        result['skip_reason'] = f"ヘッダー行が見つかりません（キー列2つ以上必要）。\n" + "\n".join(diag_lines)
        print(f"  ✗ ヘッダー未検出: {sheet_name}")
        print(f"    上位候補:")
        for dl in diag_lines[:3]:
            print(f"      {dl}")
        return result

    # 連続するヘッダー候補をフィルタ（より多いkey_countを優先）
    # 同じ行、または連続行のヘッダーは重複除去
    filtered_headers = []
    header_candidates.sort(key=lambda x: x['row'])

    for candidate in header_candidates:
        # 既存のヘッダーと重複チェック（行番号が近いか）
        is_duplicate = False
        for i, existing in enumerate(filtered_headers):
            existing_rows = set(existing['rows'])
            candidate_rows = set(candidate['rows'])
            # 行が重複するか、連続している場合
            if existing_rows & candidate_rows or max(existing_rows) + 1 >= min(candidate_rows):
                # より多くのキー列を持つ方を採用
                if candidate['key_count'] > existing['key_count'] or (
                    candidate['key_count'] == existing['key_count'] and
                    candidate['total_count'] > existing['total_count']
                ) or (
                    candidate['key_count'] == existing['key_count'] and
                    candidate['total_count'] == existing['total_count'] and
                    not candidate['is_two_row'] and existing['is_two_row']  # 単一行優先
                ):
                    filtered_headers[i] = candidate
                is_duplicate = True
                break
        if not is_duplicate:
            filtered_headers.append(candidate)

    print(f"  検出ヘッダー数: {len(filtered_headers)}")
    for h in filtered_headers:
        print(f"    行{h['rows']}: キー{h['key_count']}個 {list(h['map'].keys())}")

    # 各セクションからデータを抽出
    all_lines = []

    for section_idx, header in enumerate(filtered_headers):
        # 次のヘッダーの行番号を取得（終端判定用）
        if section_idx + 1 < len(filtered_headers):
            next_header_row = filtered_headers[section_idx + 1]['row']
        else:
            next_header_row = 1000  # 最後のセクションは1000行まで

        header_row = max(header['rows'])  # 2段ヘッダーの場合は下の行
        header_map = header['map']

        # 必須カラムチェック緩和: 名称なしでも qty+unit_price or qty+amount でOK
        has_valid_columns = (
            'name' in header_map or
            'amount' in header_map or
            ('qty' in header_map and 'unit_price' in header_map)
        )
        if not has_valid_columns:
            print(f"    セクション{section_idx+1}スキップ: 有効カラムなし {list(header_map.keys())}")
            continue

        section_lines = []
        empty_row_count = 0
        # 次のヘッダーがある場合は大きめの許容値、最終セクションは小さめ
        max_empty_rows = 20 if section_idx + 1 < len(filtered_headers) else 10

        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, max_row=next_header_row - 1), start=header_row + 1):
            cells = list(row)

            # 次のヘッダー行に達したら終了
            if row_idx >= next_header_row:
                break

            # 名称を取得
            name = ''
            if 'name' in header_map:
                name_idx = header_map['name']
                if name_idx < len(cells):
                    name = str(cells[name_idx].value or '').strip()

            # 金額を取得（先に取得して終端判定に使用）
            amount = None
            if 'amount' in header_map:
                amount_idx = header_map['amount']
                if amount_idx < len(cells):
                    amount = normalize_number(cells[amount_idx].value)

            # 単価も確認（金額がなくても単価があればデータ行の可能性）
            unit_price_val = None
            if 'unit_price' in header_map:
                up_idx = header_map['unit_price']
                if up_idx < len(cells):
                    unit_price_val = normalize_number(cells[up_idx].value)

            # 空行判定（名称・金額・単価いずれもない場合）
            if not name and not amount and not unit_price_val:
                empty_row_count += 1
                if empty_row_count >= max_empty_rows:
                    break
                continue
            else:
                empty_row_count = 0

            # 合計行や小計行をスキップ
            # 名称を正規化（空白除去）してチェック
            name_normalized = name.replace(' ', '').replace('　', '')
            skip_exact = ['計', '小計', '合計', '総合計', '工事費計', '税込合計', '税抜合計',
                          '本工事費計', '直接工事費計', '諸経費計', '一般管理費', '消費税',
                          '内訳明細書', '明細書']
            if name_normalized in skip_exact:
                continue
            # 「○○計」パターン（ただし6文字以下の短い計のみ）
            if name_normalized.endswith('計') and len(name_normalized) <= 7:
                continue
            # 「○○総計」パターン
            if '総計' in name_normalized:
                continue
            # 消費税行をスキップ
            if '消費税' in name or '税込' in name and '計' in name:
                continue
            # 注意書き行をスキップ（※で始まる）
            if name.startswith('※') or name.startswith('＊'):
                continue
            # ヘッダータイトル行をスキップ（金額・数量・単価が全てない場合の「○○書」）
            if name_normalized.endswith('書') and not amount and not unit_price_val:
                continue
            # カテゴリヘッダー行をスキップ（「○○工事」で金額・単価がない）
            if (name_normalized.endswith('工事') or name_normalized.endswith('工')) and not amount and not unit_price_val:
                continue

            # 各カラムを取得
            breakdown = ''
            if 'breakdown' in header_map:
                bd_idx = header_map['breakdown']
                if bd_idx < len(cells):
                    breakdown = str(cells[bd_idx].value or '').strip()

            qty = None
            qty_raw = None
            if 'qty' in header_map:
                qty_idx = header_map['qty']
                if qty_idx < len(cells):
                    qty_raw = cells[qty_idx].value
                    qty = normalize_number(qty_raw)

            unit = ''
            if 'unit' in header_map:
                unit_idx = header_map['unit']
                if unit_idx < len(cells):
                    unit = str(cells[unit_idx].value or '').strip()

            # "一式"/"1式" パターン処理（qty列）
            qty_raw_str = str(qty_raw or '').strip() if qty_raw else ''
            if '式' in qty_raw_str or qty_raw_str in ['一式', '1式', '１式']:
                qty = 1.0
                if not unit:
                    unit = '式'

            unit_price = None
            if 'unit_price' in header_map:
                up_idx = header_map['unit_price']
                if up_idx < len(cells):
                    unit_price = normalize_number(cells[up_idx].value)

            note = ''
            if 'note' in header_map:
                note_idx = header_map['note']
                if note_idx < len(cells):
                    note = str(cells[note_idx].value or '').strip()

            # "一式" パターン処理（name, breakdown, unit列もチェック）
            # qty が空で、他の列に "一式" があれば qty=1, unit='式'
            if qty is None:
                check_texts = [name, breakdown, unit]
                for txt in check_texts:
                    if txt and ('一式' in txt or '1式' in txt or '１式' in txt):
                        qty = 1.0
                        if not unit or unit in ['一式', '1式', '１式']:
                            unit = '式'
                        break

            # 自動計算（逆算）ロジック強化
            # 1. amount が空で qty と unit_price があれば計算
            if amount is None and qty is not None and unit_price is not None:
                amount = round(qty * unit_price, 0)

            # 2. qty が空で amount と unit_price があれば逆算
            if qty is None and amount is not None and unit_price is not None and unit_price > 0:
                qty = round(amount / unit_price, 2)  # 小数第2位まで

            # 3. unit_price が空で amount と qty があれば逆算
            if unit_price is None and amount is not None and qty is not None and qty > 0:
                unit_price = round(amount / qty, 0)  # 単価は整数

            # 4. amount だけある場合、qty=1 と仮定して unit_price を逆算
            if qty is None and unit_price is None and amount is not None and amount > 0:
                qty = 1.0
                unit_price = amount
                if not unit:
                    unit = '式'

            # 最低限nameかamountがあれば行を追加
            if name or (amount is not None and amount != 0):
                section_lines.append({
                    'sheet_name': sheet_name,
                    'row_no': row_idx,
                    'name': name or '（名称なし）',
                    'breakdown': breakdown,
                    'qty': qty,
                    'unit': unit,
                    'unit_price': unit_price,
                    'amount': amount,
                    'note': note,
                })

        print(f"    セクション{section_idx+1}(行{header['rows']}): {len(section_lines)}行")
        result['sections'].append({
            'header_rows': header['rows'],
            'header_map': list(header['map'].keys()),
            'line_count': len(section_lines)
        })
        all_lines.extend(section_lines)

    # 結果を設定
    result['lines'] = all_lines
    if filtered_headers:
        # 最初のヘッダーを代表として設定（後方互換）
        result['header_row'] = filtered_headers[0]['row']
        result['header_rows'] = filtered_headers[0]['rows']
        result['header_map'] = filtered_headers[0]['map']

    if not result['lines']:
        result['skip_reason'] = 'データ行が見つかりません（ヘッダー行の下にデータがありません）'
        print(f"  ✗ データ行なし: {sheet_name}")
    else:
        print(f"  ✓ {len(result['lines'])}行のデータを抽出（{len(filtered_headers)}セクション）: {sheet_name}")

    return result



def classify_cost_category(name: str) -> str:
    """
    名称から原価カテゴリを推定
    """
    name_lower = name.lower()

    if any(kw in name_lower for kw in ['労務', '人工', '作業員', '技術者', '人件', '日当']):
        return 'labor'
    elif any(kw in name_lower for kw in ['外注', '下請', '業者', '協力']):
        return 'subcontract'
    elif any(kw in name_lower for kw in ['材料', '資材', '管', 'パイプ', 'コンクリート', 'アスファルト', '鋼材']):
        return 'material'
    elif any(kw in name_lower for kw in ['機械', '重機', 'リース', 'クレーン', 'バックホウ', 'ダンプ']):
        return 'machine'
    else:
        return 'expense'


@app.post("/api/projects/{project_id}/imports/estimate")
async def import_estimate(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    見積/予算/原価Excelをアップロードしてプレビュー生成
    ドラフト状態のインポートレコードを作成し、import_idを返す
    """
    # プロジェクト存在確認
    project = db.query(ProjectModel).filter_by(id=project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    try:
        # ファイル内容を読み取り
        content = await file.read()
        file_hash = hashlib.sha256(content).hexdigest()

        # ファイル保存
        filename = f"{uuid.uuid4()}_{file.filename}"
        file_path = UPLOAD_DIR / filename
        with open(file_path, "wb") as f:
            f.write(content)

        # Excel解析（新形式: dictを返す）
        parse_result = parse_excel_to_lines(file_path)
        parsed_lines = parse_result['lines']

        # 0件の場合は詳細な原因を返す
        if not parsed_lines:
            error_reasons = []
            if parse_result['parse_errors']:
                error_reasons.extend(parse_result['parse_errors'])
            if parse_result['sheets_skipped']:
                for skip in parse_result['sheets_skipped']:
                    reason_msg = f"シート「{skip['name']}」: {skip['reason']}"
                    # 候補行があれば追記
                    if skip.get('candidates'):
                        top_cands = skip['candidates'][:2]
                        cand_str = ', '.join([f"行{c['row']}({','.join(c['columns'][:3])})" for c in top_cands])
                        reason_msg += f" [候補: {cand_str}]"
                    error_reasons.append(reason_msg)
            if not error_reasons:
                error_reasons.append('ヘッダー行（名称/数量/単価/金額が2つ以上揃う行）を検出できませんでした')

            # 理由コードを日本語に変換
            reason_labels = {
                'parse_error': 'Excelファイルの読み込みエラー',
                'empty_workbook': '空のワークブック',
                'header_not_found': 'ヘッダー行が見つかりません',
                'required_columns_missing': '必須カラム（名称/金額）がありません',
                'no_data_rows': 'データ行がありません',
            }
            reason_code = parse_result.get('reason', 'unknown')

            return {
                'status': 'warning',
                'import_id': None,
                'filename': file.filename,
                'preview': {
                    'lines': [],
                    'total_amount': 0,
                    'line_count': 0,
                    'sheets_processed': [],
                    'sheets_skipped': parse_result['sheets_skipped'],
                    'missing_columns': parse_result['missing_columns'],
                    'error_reasons': error_reasons,
                    'reason': reason_code,
                    'reason_label': reason_labels.get(reason_code, reason_code),
                }
            }

        # EstimateImportレコード作成（draft状態）
        estimate_import = EstimateImportModel(
            project_id=project_id,
            original_filename=file.filename,
            storage_path=str(file_path),
            file_hash=file_hash,
            meta_json=json.dumps({
                'sheets_processed': parse_result['sheets_processed'],
                'sheets_skipped': parse_result['sheets_skipped'],
                'detected_headers': {k: v for k, v in parse_result['detected_headers'].items()},
                'line_count': len(parsed_lines)
            }),
            status='draft'
        )
        db.add(estimate_import)
        db.flush()  # IDを取得するため

        # EstimateLineレコード作成
        for line in parsed_lines:
            category = classify_cost_category(line['name'])
            estimate_line = EstimateLineModel(
                import_id=estimate_import.id,
                sheet_name=line['sheet_name'],
                row_no=line['row_no'],
                kind='estimate',  # デフォルトは見積
                name=line['name'],
                breakdown=line['breakdown'],
                qty=line['qty'],
                unit=line['unit'],
                unit_price=line['unit_price'],
                amount=line['amount'],
                note=line['note'],
                category=category
            )
            db.add(estimate_line)

        db.commit()

        # プレビューデータ作成
        preview_lines = []
        for line in parsed_lines:
            preview_lines.append({
                'sheet_name': line['sheet_name'],
                'row_no': line['row_no'],
                'name': line['name'],
                'breakdown': line['breakdown'],
                'qty': line['qty'],
                'unit': line['unit'],
                'unit_price': line['unit_price'],
                'amount': line['amount'],
                'note': line['note'],
                'category': classify_cost_category(line['name'])
            })

        return {
            'status': 'success',
            'import_id': estimate_import.id,
            'filename': file.filename,
            'preview': {
                'lines': preview_lines,
                'total_amount': sum(l['amount'] or 0 for l in parsed_lines),
                'line_count': len(parsed_lines),
                'sheets_processed': parse_result['sheets_processed'],
                'sheets_skipped': parse_result['sheets_skipped'],
                'missing_columns': parse_result['missing_columns'],
                'detected_headers': parse_result.get('detected_headers', {}),
                'value_stats': parse_result.get('value_stats', {}),
            }
        }

    except Exception as e:
        db.rollback()
        import traceback
        print(f"Import error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"インポートエラー: {str(e)}")


class CommitRequest(BaseModel):
    """コミットリクエスト"""
    kind: Optional[str] = 'estimate'  # estimate, budget, actual
    month: Optional[str] = None  # YYYY-MM (budget/actual用)


@app.post("/api/imports/{import_id}/commit")
async def commit_import(
    import_id: str,
    request: CommitRequest = None,
    db: Session = Depends(get_db)
):
    """
    ドラフト状態のインポートを確定
    kind=actualの場合はcost_recordsへ同期
    """
    # インポートレコード取得
    estimate_import = db.query(EstimateImportModel).filter_by(id=import_id).first()
    if not estimate_import:
        raise HTTPException(status_code=404, detail="インポートが見つかりません")

    if estimate_import.status != 'draft':
        raise HTTPException(status_code=400, detail="既にコミット済みです")

    try:
        # kindとmonthを更新
        kind = request.kind if request else 'estimate'
        month = request.month if request else None

        lines = db.query(EstimateLineModel).filter_by(import_id=import_id).all()

        for line in lines:
            line.kind = kind
            if month:
                line.month = month

        # kind=actualの場合はcost_recordsへ同期
        if kind == 'actual':
            for line in lines:
                cost_record = CostRecordModel(
                    project_id=estimate_import.project_id,
                    category=line.category or 'expense',
                    item_name=line.name,
                    quantity=line.qty,
                    unit=line.unit,
                    unit_price=line.unit_price,
                    amount=line.amount
                )
                db.add(cost_record)

        # ステータス更新
        estimate_import.status = 'committed'
        db.commit()

        return {
            'status': 'success',
            'message': '保存しました',
            'import_id': import_id,
            'kind': kind,
            'line_count': len(lines)
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"コミットエラー: {str(e)}")


@app.get("/api/projects/{project_id}/estimate-lines")
async def get_estimate_lines(
    project_id: str,
    year: Optional[int] = None,
    kind: Optional[str] = None,
    month: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    プロジェクトの見積明細を取得（年度・種類・月フィルタ対応）
    month: YYYY-MM形式（予算・原価の場合に使用）
    """
    # インポート一覧を取得
    query = db.query(EstimateImportModel).filter_by(
        project_id=project_id,
        status='committed'
    )
    imports = query.all()

    result_lines = []
    for imp in imports:
        lines_query = db.query(EstimateLineModel).filter_by(import_id=imp.id)

        if kind:
            lines_query = lines_query.filter_by(kind=kind)

        # 月フィルタ（DBレベル）
        if month:
            lines_query = lines_query.filter(EstimateLineModel.month == month)

        # sort_orderで並び替え
        lines_query = lines_query.order_by(EstimateLineModel.sort_order, EstimateLineModel.row_no)

        lines = lines_query.all()

        for line in lines:
            # 年度フィルタ
            if year:
                if line.month:
                    line_year = int(line.month.split('-')[0])
                    if line_year != year:
                        continue
                else:
                    # monthがない場合はインポート日時でフィルタ
                    if imp.uploaded_at.year != year:
                        continue

            result_lines.append({
                'id': line.id,
                'import_id': line.import_id,
                'sheet_name': line.sheet_name,
                'row_no': line.row_no,
                'kind': line.kind,
                'name': line.name,
                'breakdown': line.breakdown,
                'qty': line.qty,
                'unit': line.unit,
                'unit_price': line.unit_price,
                'amount': line.amount,
                'note': line.note,
                'category': line.category,
                'month': line.month,
                'sort_order': line.sort_order or 0,
                'created_at': line.created_at.isoformat() if line.created_at else None
            })

    # sort_orderとrow_noで最終ソート
    result_lines.sort(key=lambda x: (x['sort_order'], x['row_no'] or 0))

    return {
        'status': 'success',
        'project_id': project_id,
        'lines': result_lines,
        'total_amount': sum(l['amount'] or 0 for l in result_lines),
        'filters': {
            'year': year,
            'kind': kind,
            'month': month
        }
    }


class ReorderRequest(BaseModel):
    line_ids: List[str]
    sort_orders: List[int]


@app.patch("/api/projects/{project_id}/estimate-lines/reorder")
async def reorder_estimate_lines(
    project_id: str,
    request: ReorderRequest,
    db: Session = Depends(get_db)
):
    """
    見積明細の並び順を更新
    """
    if len(request.line_ids) != len(request.sort_orders):
        raise HTTPException(status_code=400, detail="line_idsとsort_ordersの長さが一致しません")

    try:
        updated_count = 0
        for line_id, sort_order in zip(request.line_ids, request.sort_orders):
            line = db.query(EstimateLineModel).filter_by(id=line_id).first()
            if line:
                # プロジェクトの所有確認
                imp = db.query(EstimateImportModel).filter_by(id=line.import_id).first()
                if imp and imp.project_id == project_id:
                    line.sort_order = sort_order
                    updated_count += 1

        db.commit()

        return {
            'status': 'success',
            'message': f'{updated_count}件の並び順を更新しました',
            'updated_count': updated_count
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"更新エラー: {str(e)}")


@app.get("/api/projects/{project_id}/attachments")
async def get_attachments(
    project_id: str,
    db: Session = Depends(get_db)
):
    """
    プロジェクトの添付ファイル一覧を取得
    """
    attachments = db.query(AttachmentModel).filter_by(project_id=project_id).all()

    return {
        'status': 'success',
        'project_id': project_id,
        'attachments': [
            {
                'id': att.id,
                'import_id': att.import_id,
                'type': att.type,
                'filename': att.filename,
                'storage_path': att.storage_path,
                'uploaded_at': att.uploaded_at.isoformat() if att.uploaded_at else None,
                'download_url': f"/uploads/{Path(att.storage_path).name}" if att.storage_path else None
            }
            for att in attachments
        ]
    }


@app.post("/api/projects/{project_id}/attachments")
async def upload_attachment(
    project_id: str,
    file: UploadFile = File(...),
    type: str = Form('other'),
    import_id: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    添付ファイルをアップロード
    """
    # ファイル保存
    filename = f"{uuid.uuid4()}_{file.filename}"
    file_path = UPLOAD_DIR / filename
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    # DBレコード作成
    attachment = AttachmentModel(
        project_id=project_id,
        import_id=import_id,
        type=type,
        filename=file.filename,
        storage_path=str(file_path)
    )
    db.add(attachment)
    db.commit()

    return {
        'status': 'success',
        'attachment_id': attachment.id,
        'filename': file.filename
    }


@app.post("/api/budget/create")
async def create_budget(budget: Budget):
    """
    実行予算を作成
    """
    try:
        # データベースに保存（デモ用）
        budgets_db[budget.id] = budget
        
        return {
            "status": "success",
            "budget_id": budget.id,
            "message": "実行予算を作成しました"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"作成エラー: {str(e)}")

@app.get("/api/budget/{budget_id}")
async def get_budget(budget_id: str):
    """
    実行予算を取得
    """
    if budget_id not in budgets_db:
        raise HTTPException(status_code=404, detail="予算が見つかりません")
    
    return budgets_db[budget_id].dict()

@app.post("/api/budget/{budget_id}/pdf")
async def generate_pdf(budget_id: str):
    """
    見積書PDFを生成
    """
    if budget_id not in budgets_db:
        raise HTTPException(status_code=404, detail="予算が見つかりません")
    
    try:
        budget = budgets_db[budget_id]
        pdf_path = generate_estimate_pdf(budget)
        
        return FileResponse(
            path=str(pdf_path),
            filename=f"estimate_{budget_id}.pdf",
            media_type="application/pdf"
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF生成エラー: {str(e)}")

@app.get("/api/projects")
async def get_projects(db: Session = Depends(get_db)):
    """
    工事一覧を取得（DBから）
    """
    projects = db.query(ProjectModel).order_by(ProjectModel.created_at.desc()).all()

    return {
        "status": "success",
        "projects": [
            {
                "id": p.id,
                "name": p.name,
                "client": p.client_name,
                "contract_amount": p.contract_amount or 0,
                "budget_amount": p.budget_amount or 0,
                "actual_cost": p.actual_cost or 0,
                "profit_rate": p.profit_rate or 0,
                "progress": p.progress or 0,
                "status": p.status or "active",
                "created_at": p.created_at.isoformat() if p.created_at else None
            }
            for p in projects
        ]
    }

@app.post("/api/projects")
async def create_project(project: ProjectCreate, db: Session = Depends(get_db)):
    """
    工事プロジェクトを作成
    """
    new_project = ProjectModel(
        id=str(uuid.uuid4()),
        name=project.name,
        client_name=project.client_name,
        contract_amount=project.contract_amount,
        budget_amount=project.budget_amount,
        status=project.status,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow()
    )

    db.add(new_project)
    db.commit()
    db.refresh(new_project)

    return {
        "status": "success",
        "project": {
            "id": new_project.id,
            "name": new_project.name,
            "client": new_project.client_name,
            "contract_amount": new_project.contract_amount,
            "budget_amount": new_project.budget_amount,
            "status": new_project.status,
            "created_at": new_project.created_at.isoformat()
        }
    }

# =====================================
# 見積API（Project紐づけ）
# =====================================

@app.get("/api/projects/{project_id}/estimates")
async def get_project_estimates(project_id: str, db: Session = Depends(get_db)):
    """
    プロジェクトに紐づく見積一覧を取得
    """
    # プロジェクト存在確認
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    estimates = db.query(EstimateModel).filter(EstimateModel.project_id == project_id).all()

    return {
        "status": "success",
        "project_id": project_id,
        "estimates": [
            {
                "id": e.id,
                "name": e.file_name or "見積",
                "amount": e.total_amount or 0,
                "estimate_status": e.status or "draft",
                "created_at": e.created_at.isoformat() if e.created_at else None
            }
            for e in estimates
        ]
    }

@app.post("/api/projects/{project_id}/estimates")
async def create_project_estimate(project_id: str, estimate: EstimateCreate, db: Session = Depends(get_db)):
    """
    プロジェクトに見積を追加
    """
    # プロジェクト存在確認
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    new_estimate = EstimateModel(
        id=str(uuid.uuid4()),
        project_id=project_id,
        file_name=estimate.name,
        total_amount=estimate.amount,
        created_at=datetime.utcnow()
    )

    db.add(new_estimate)
    db.commit()
    db.refresh(new_estimate)

    return {
        "status": "success",
        "estimate": {
            "id": new_estimate.id,
            "name": new_estimate.file_name,
            "amount": new_estimate.total_amount,
            "estimate_status": new_estimate.status or "draft",
            "project_id": new_estimate.project_id,
            "created_at": new_estimate.created_at.isoformat()
        }
    }

@app.post("/api/estimates/{estimate_id}/order")
async def order_estimate(estimate_id: str, db: Session = Depends(get_db)):
    """
    見積を受注に変更
    """
    estimate = db.query(EstimateModel).filter(EstimateModel.id == estimate_id).first()
    if not estimate:
        raise HTTPException(status_code=404, detail="見積が見つかりません")

    if estimate.status == "ordered":
        raise HTTPException(status_code=400, detail="既に受注済みです")

    estimate.status = "ordered"
    db.commit()
    db.refresh(estimate)

    return {
        "status": "success",
        "message": "受注処理が完了しました",
        "estimate": {
            "id": estimate.id,
            "name": estimate.file_name,
            "amount": estimate.total_amount,
            "estimate_status": estimate.status,
            "project_id": estimate.project_id
        }
    }

# =====================================
# 原価API
# =====================================

@app.get("/api/projects/{project_id}/costs")
async def get_project_costs(project_id: str, month: Optional[str] = None, db: Session = Depends(get_db)):
    """
    プロジェクトの原価一覧を取得
    month: YYYY-MM形式で月絞り込み（オプション）
    """
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    query = db.query(CostRecordModel).filter(CostRecordModel.project_id == project_id)

    # 月フィルター
    if month:
        try:
            year, mon = month.split('-')
            from sqlalchemy import extract
            query = query.filter(
                extract('year', CostRecordModel.created_at) == int(year),
                extract('month', CostRecordModel.created_at) == int(mon)
            )
        except:
            pass

    costs = query.order_by(CostRecordModel.created_at.desc()).all()

    # カテゴリ別小計
    category_subtotals = {}
    for c in costs:
        cat = c.category or 'other'
        category_subtotals[cat] = category_subtotals.get(cat, 0) + (c.amount or 0)

    return {
        "status": "success",
        "project_id": project_id,
        "month": month,
        "costs": [
            {
                "id": c.id,
                "category": c.category,
                "amount": c.amount or 0,
                "note": c.item_name or "",
                "cost_date": c.created_at.isoformat() if c.created_at else None
            }
            for c in costs
        ],
        "category_subtotals": category_subtotals,
        "total": sum(c.amount or 0 for c in costs)
    }

@app.post("/api/projects/{project_id}/costs")
async def create_project_cost(project_id: str, cost: CostCreate, db: Session = Depends(get_db)):
    """
    プロジェクトに原価を追加
    """
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    new_cost = CostRecordModel(
        id=str(uuid.uuid4()),
        project_id=project_id,
        category=cost.category,
        amount=cost.amount,
        item_name=cost.note,
        created_at=datetime.utcnow()
    )

    db.add(new_cost)
    db.commit()
    db.refresh(new_cost)

    return {
        "status": "success",
        "cost": {
            "id": new_cost.id,
            "category": new_cost.category,
            "amount": new_cost.amount,
            "note": new_cost.item_name,
            "project_id": new_cost.project_id,
            "cost_date": new_cost.created_at.isoformat()
        }
    }


@app.put("/api/costs/{cost_id}")
async def update_cost(cost_id: str, cost: CostCreate, db: Session = Depends(get_db)):
    """
    原価を更新
    """
    existing = db.query(CostRecordModel).filter(CostRecordModel.id == cost_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="原価が見つかりません")

    existing.category = cost.category
    existing.amount = cost.amount
    existing.item_name = cost.note

    db.commit()
    db.refresh(existing)

    return {
        "status": "success",
        "cost": {
            "id": existing.id,
            "category": existing.category,
            "amount": existing.amount,
            "note": existing.item_name,
            "project_id": existing.project_id,
            "cost_date": existing.created_at.isoformat() if existing.created_at else None
        }
    }


@app.delete("/api/costs/{cost_id}")
async def delete_cost(cost_id: str, db: Session = Depends(get_db)):
    """
    原価を削除
    """
    existing = db.query(CostRecordModel).filter(CostRecordModel.id == cost_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="原価が見つかりません")

    db.delete(existing)
    db.commit()

    return {"status": "success", "message": "削除しました"}


@app.get("/api/projects/{project_id}/summary")
async def get_project_summary(project_id: str, db: Session = Depends(get_db)):
    """
    プロジェクトのサマリー（売上・原価・粗利・請求・入金）
    売上 = 請求合計（issued + paid）
    """
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    # 請求・入金集計
    invoices = db.query(InvoiceModel).filter(InvoiceModel.project_id == project_id).all()
    billed_total = sum(i.amount or 0 for i in invoices if i.status in ["issued", "paid"])
    paid_total = sum(i.amount or 0 for i in invoices if i.status == "paid")
    unpaid_amount = billed_total - paid_total

    # 売上 = 請求合計（issued + paid）
    revenue = billed_total

    # 原価合計
    costs = db.query(CostRecordModel).filter(CostRecordModel.project_id == project_id).all()
    cost_total = sum(c.amount or 0 for c in costs)

    # 粗利
    gross_profit = revenue - cost_total
    gross_margin = (gross_profit / revenue * 100) if revenue > 0 else 0

    return {
        "status": "success",
        "project_id": project_id,
        "summary": {
            "revenue": revenue,
            "cost_total": cost_total,
            "gross_profit": gross_profit,
            "gross_margin": round(gross_margin, 1),
            "paid_total": paid_total,
            "unpaid_amount": unpaid_amount
        }
    }

@app.get("/api/dashboard/stats")
async def get_dashboard_stats():
    """
    ダッシュボード統計情報を取得
    """
    return {
        "active_projects": 23,
        "profit_margin": 18.5,
        "revenue": 125000000,
        "alerts": 2
    }

# =====================================
# 予算一覧API（BudgetList用）
# =====================================

@app.get("/api/budgets")
async def get_budgets():
    """
    予算一覧を取得
    """
    # デモデータを含める
    demo_budgets = [
        {
            "id": "demo-1",
            "projectName": "広島自動車道改修工事",
            "projectCode": "PRJ-2026-001",
            "customerName": "NEXCO西日本",
            "totalAmount": 15000000,
            "laborCost": 3000000,
            "outsourceCost": 5000000,
            "materialCost": 4000000,
            "equipmentCost": 2000000,
            "expenseCost": 1000000,
            "profitRate": 18.5,
            "status": "approved",
            "createdBy": "管理者",
            "createdAt": "2026-01-10",
            "updatedAt": "2026-01-10"
        },
        {
            "id": "demo-2",
            "projectName": "○○市水道管工事",
            "projectCode": "PRJ-2026-002",
            "customerName": "○○市水道局",
            "totalAmount": 8500000,
            "laborCost": 2000000,
            "outsourceCost": 3000000,
            "materialCost": 2000000,
            "equipmentCost": 1000000,
            "expenseCost": 500000,
            "profitRate": 12.5,
            "status": "pending",
            "createdBy": "田中",
            "createdAt": "2026-01-08",
            "updatedAt": "2026-01-11"
        },
        {
            "id": "demo-3",
            "projectName": "新築マンション電気工事",
            "projectCode": "PRJ-2026-003",
            "customerName": "○○建設株式会社",
            "totalAmount": 5200000,
            "laborCost": 1500000,
            "outsourceCost": 1500000,
            "materialCost": 1500000,
            "equipmentCost": 500000,
            "expenseCost": 200000,
            "profitRate": 22.0,
            "status": "draft",
            "createdBy": "佐藤",
            "createdAt": "2026-01-12",
            "updatedAt": "2026-01-12"
        },
    ]

    # DBに保存されたデータとデモデータを結合
    all_budgets = demo_budgets + list(simple_budgets_db.values())
    return all_budgets

@app.post("/api/budgets")
async def create_simple_budget(budget: SimpleBudget):
    """
    予算を作成（シンプル版）
    """
    budget_id = str(uuid.uuid4())
    now = datetime.now().isoformat()

    budget_data = {
        "id": budget_id,
        "projectName": budget.projectName,
        "projectCode": budget.projectCode,
        "customerName": budget.customerName,
        "totalAmount": budget.totalAmount,
        "laborCost": budget.laborCost,
        "outsourceCost": budget.outsourceCost,
        "materialCost": budget.materialCost,
        "equipmentCost": budget.equipmentCost,
        "expenseCost": budget.expenseCost,
        "profitRate": budget.profitRate,
        "status": budget.status,
        "createdBy": budget.createdBy,
        "createdAt": now,
        "updatedAt": now
    }

    simple_budgets_db[budget_id] = budget_data

    return budget_data

@app.delete("/api/budgets/{budget_id}")
async def delete_budget(budget_id: str):
    """
    予算を削除
    """
    if budget_id.startswith("demo-"):
        raise HTTPException(status_code=400, detail="デモデータは削除できません")

    if budget_id not in simple_budgets_db:
        raise HTTPException(status_code=404, detail="予算が見つかりません")

    del simple_budgets_db[budget_id]
    return {"status": "success", "message": "削除しました"}


# =====================================
# 日報API
# =====================================

@app.get("/api/projects/{project_id}/daily-reports")
async def get_project_daily_reports(project_id: str, db: Session = Depends(get_db)):
    """
    プロジェクトの日報一覧を取得
    """
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    reports = db.query(DailyReportModel).filter(
        DailyReportModel.project_id == project_id
    ).order_by(DailyReportModel.work_date.desc()).all()

    result = []
    for r in reports:
        items = db.query(DailyReportItemModel).filter(
            DailyReportItemModel.daily_report_id == r.id
        ).all()
        result.append({
            "id": r.id,
            "work_date": r.work_date.strftime("%Y-%m-%d") if r.work_date else None,
            "foreman_name": r.foreman_name,
            "notes": r.notes,
            "total_amount": r.total_amount or 0,
            "items": [
                {
                    "id": i.id,
                    "worker_name": i.worker_name,
                    "hours": i.hours,
                    "wage_rate": i.wage_rate,
                    "amount": i.amount
                }
                for i in items
            ],
            "created_at": r.created_at.isoformat() if r.created_at else None
        })

    return {
        "status": "success",
        "project_id": project_id,
        "daily_reports": result
    }


@app.post("/api/projects/{project_id}/daily-reports")
async def create_project_daily_report(project_id: str, report: DailyReportCreate, db: Session = Depends(get_db)):
    """
    日報を作成し、労務費を自動登録
    """
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    # 日報を作成
    report_id = str(uuid.uuid4())
    work_date = datetime.strptime(report.work_date, "%Y-%m-%d")

    new_report = DailyReportModel(
        id=report_id,
        project_id=project_id,
        work_date=work_date,
        foreman_name=report.foreman_name,
        notes=report.notes,
        created_at=datetime.utcnow()
    )
    db.add(new_report)

    # 明細を作成し、合計金額を計算
    total_amount = 0
    items_data = []
    for item in report.items:
        item_id = str(uuid.uuid4())
        amount = int(item.hours * item.wage_rate)
        total_amount += amount

        new_item = DailyReportItemModel(
            id=item_id,
            daily_report_id=report_id,
            worker_name=item.worker_name,
            hours=item.hours,
            wage_rate=item.wage_rate,
            amount=amount
        )
        db.add(new_item)
        items_data.append({
            "id": item_id,
            "worker_name": item.worker_name,
            "hours": item.hours,
            "wage_rate": item.wage_rate,
            "amount": amount
        })

    # 合計金額を更新
    new_report.total_amount = total_amount

    # 労務費を自動登録（UPSERT: daily_report_idで既存チェック）
    existing_cost = db.query(CostRecordModel).filter(
        CostRecordModel.daily_report_id == report_id
    ).first()

    if existing_cost:
        # 更新
        existing_cost.amount = total_amount
        existing_cost.created_at = work_date
    else:
        # 新規作成
        labor_cost = CostRecordModel(
            id=str(uuid.uuid4()),
            project_id=project_id,
            daily_report_id=report_id,
            category="labor",
            item_name="日報集計",
            amount=total_amount,
            created_at=work_date
        )
        db.add(labor_cost)

    db.commit()
    db.refresh(new_report)

    return {
        "status": "success",
        "daily_report": {
            "id": new_report.id,
            "work_date": new_report.work_date.strftime("%Y-%m-%d"),
            "foreman_name": new_report.foreman_name,
            "notes": new_report.notes,
            "total_amount": total_amount,
            "items": items_data,
            "created_at": new_report.created_at.isoformat()
        }
    }


# =====================================
# 請求API
# =====================================

@app.get("/api/projects/{project_id}/invoices")
async def get_project_invoices(project_id: str, db: Session = Depends(get_db)):
    """
    プロジェクトの請求一覧を取得
    """
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    invoices = db.query(InvoiceModel).filter(
        InvoiceModel.project_id == project_id
    ).order_by(InvoiceModel.billing_month.desc()).all()

    return {
        "status": "success",
        "project_id": project_id,
        "invoices": [
            {
                "id": i.id,
                "billing_month": i.billing_month,
                "amount": i.amount or 0,
                "status": i.status or "draft",
                "notes": i.notes,
                "issued_at": i.issued_at.isoformat() if i.issued_at else None,
                "paid_at": i.paid_at.isoformat() if i.paid_at else None,
                "created_at": i.created_at.isoformat() if i.created_at else None
            }
            for i in invoices
        ]
    }


@app.post("/api/projects/{project_id}/invoices")
async def create_project_invoice(project_id: str, invoice: InvoiceCreate, db: Session = Depends(get_db)):
    """
    請求を作成
    """
    project = db.query(ProjectModel).filter(ProjectModel.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="プロジェクトが見つかりません")

    new_invoice = InvoiceModel(
        id=str(uuid.uuid4()),
        project_id=project_id,
        billing_month=invoice.billing_month,
        amount=invoice.amount,
        notes=invoice.notes,
        status="draft",
        created_at=datetime.utcnow()
    )

    db.add(new_invoice)
    db.commit()
    db.refresh(new_invoice)

    return {
        "status": "success",
        "invoice": {
            "id": new_invoice.id,
            "billing_month": new_invoice.billing_month,
            "amount": new_invoice.amount,
            "status": new_invoice.status,
            "notes": new_invoice.notes,
            "created_at": new_invoice.created_at.isoformat()
        }
    }


@app.patch("/api/invoices/{invoice_id}/status")
async def update_invoice_status(invoice_id: str, update: InvoiceStatusUpdate, db: Session = Depends(get_db)):
    """
    請求ステータスを更新（draft→issued→paid）
    """
    invoice = db.query(InvoiceModel).filter(InvoiceModel.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="請求が見つかりません")

    valid_statuses = ["draft", "issued", "paid"]
    if update.status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"無効なステータス: {update.status}")

    invoice.status = update.status
    if update.status == "issued":
        invoice.issued_at = datetime.utcnow()
    elif update.status == "paid":
        invoice.paid_at = datetime.utcnow()

    db.commit()
    db.refresh(invoice)

    return {
        "status": "success",
        "invoice": {
            "id": invoice.id,
            "billing_month": invoice.billing_month,
            "amount": invoice.amount,
            "status": invoice.status,
            "issued_at": invoice.issued_at.isoformat() if invoice.issued_at else None,
            "paid_at": invoice.paid_at.isoformat() if invoice.paid_at else None
        }
    }


# =====================================
# LINE通知API
# =====================================

def send_line_push(user_id: str, message: str) -> dict:
    """
    LINE Messaging APIでプッシュメッセージを送信
    """
    token = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
    if not token:
        return {"success": False, "error": "LINE_CHANNEL_ACCESS_TOKEN not configured"}

    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    payload = {
        "to": user_id,
        "messages": [
            {
                "type": "text",
                "text": message
            }
        ]
    }

    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            return {"success": True}
        else:
            return {"success": False, "error": response.text, "status_code": response.status_code}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/line/webhook")
async def line_webhook(request: dict):
    """
    LINE Webhookエンドポイント（User ID取得用）
    """
    events = request.get("events", [])
    for event in events:
        user_id = event.get("source", {}).get("userId")
        if user_id:
            print(f"=== LINE User ID: {user_id} ===")
            return {"status": "ok", "user_id": user_id}
    return {"status": "ok"}


@app.post("/api/notify/test")
async def send_test_notification():
    """
    テスト通知を送信（管理者のLINEに）
    """
    admin_user_id = os.getenv("LINE_ADMIN_USER_ID")
    if not admin_user_id:
        raise HTTPException(status_code=500, detail="LINE_ADMIN_USER_ID not configured")

    message = "【sunyuTECH-DX】テスト通知OK"
    result = send_line_push(admin_user_id, message)

    if result.get("success"):
        return {
            "status": "success",
            "message": "テスト通知を送信しました"
        }
    else:
        raise HTTPException(
            status_code=500,
            detail=f"LINE送信エラー: {result.get('error')}"
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
