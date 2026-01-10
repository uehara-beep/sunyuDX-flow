"""
Excel出力モジュール
実行予算書・工事台帳をExcel出力
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime

class ExcelGenerator:
    """Excel出力クラス"""
    
    # スタイル定義
    HEADER_FILL = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
    HEADER_FONT = Font(name='游ゴシック', size=11, bold=True, color="FFFFFF")
    CATEGORY_FILL = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
    CATEGORY_FONT = Font(name='游ゴシック', size=10, bold=True, color="FFFFFF")
    NORMAL_FONT = Font(name='游ゴシック', size=10)
    TOTAL_FILL = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    TOTAL_FONT = Font(name='游ゴシック', size=11, bold=True, color="FFFFFF")
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def generate_budget_excel(budget_data: Dict[str, Any], output_path: Path) -> Path:
        """
        実行予算書をExcel出力
        
        Args:
            budget_data: 予算データ
            output_path: 出力先パス
        
        Returns:
            出力ファイルのパス
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "実行予算書"
        
        # 列幅設定
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        row = 1
        
        # タイトル
        ws.merge_cells(f'A{row}:E{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "実行予算書"
        title_cell.font = Font(name='游ゴシック', size=18, bold=True, color="0066cc")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
        row += 2
        
        # プロジェクト情報
        ws[f'A{row}'] = "工事名:"
        ws[f'B{row}'] = budget_data.get('project_name', '')
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'A{row}'].font = Font(name='游ゴシック', size=10, bold=True)
        row += 1
        
        ws[f'A{row}'] = "発注者:"
        ws[f'B{row}'] = budget_data.get('client_name', '')
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'A{row}'].font = Font(name='游ゴシック', size=10, bold=True)
        row += 1
        
        ws[f'A{row}'] = "作成日:"
        ws[f'B{row}'] = datetime.now().strftime('%Y年%m月%d日')
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'A{row}'].font = Font(name='游ゴシック', size=10, bold=True)
        row += 2
        
        # ヘッダー
        headers = ['項目', '数量', '単位', '単価', '金額']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row, col_idx)
            cell.value = header
            cell.font = ExcelGenerator.HEADER_FONT
            cell.fill = ExcelGenerator.HEADER_FILL
            cell.border = ExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # 5科目データ
        categories = [
            ('材料費', budget_data.get('material', {})),
            ('労務費', budget_data.get('labor', {})),
            ('機械費', budget_data.get('equipment', {})),
            ('外注費', budget_data.get('subcontract', {})),
            ('経費', budget_data.get('expense', {})),
        ]
        
        for category_name, category_data in categories:
            items = category_data.get('items', [])
            category_total = category_data.get('total', 0)
            
            if category_total > 0:
                # カテゴリーヘッダー
                ws.merge_cells(f'A{row}:E{row}')
                cat_cell = ws[f'A{row}']
                cat_cell.value = category_name
                cat_cell.font = ExcelGenerator.CATEGORY_FONT
                cat_cell.fill = ExcelGenerator.CATEGORY_FILL
                cat_cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
                
                # 項目
                for item in items:
                    ws[f'A{row}'] = item.get('name', '')
                    ws[f'B{row}'] = item.get('quantity', 0)
                    ws[f'C{row}'] = item.get('unit', '')
                    ws[f'D{row}'] = item.get('unit_price', 0)
                    ws[f'E{row}'] = item.get('amount', 0)
                    
                    # 数値フォーマット
                    ws[f'B{row}'].number_format = '#,##0.0'
                    ws[f'D{row}'].number_format = '¥#,##0'
                    ws[f'E{row}'].number_format = '¥#,##0'
                    
                    # 右寄せ
                    ws[f'B{row}'].alignment = Alignment(horizontal='right')
                    ws[f'D{row}'].alignment = Alignment(horizontal='right')
                    ws[f'E{row}'].alignment = Alignment(horizontal='right')
                    
                    # 罫線
                    for col_idx in range(1, 6):
                        ws.cell(row, col_idx).border = ExcelGenerator.THIN_BORDER
                    
                    row += 1
                
                # カテゴリー小計
                ws[f'A{row}'] = f"{category_name} 小計"
                ws[f'E{row}'] = category_total
                ws[f'A{row}'].font = Font(name='游ゴシック', size=10, bold=True)
                ws[f'E{row}'].font = Font(name='游ゴシック', size=10, bold=True)
                ws[f'E{row}'].number_format = '¥#,##0'
                ws[f'E{row}'].alignment = Alignment(horizontal='right')
                
                for col_idx in range(1, 6):
                    ws.cell(row, col_idx).border = ExcelGenerator.THIN_BORDER
                
                row += 2
        
        # 合計
        ws[f'A{row}'] = "実行予算合計"
        ws[f'E{row}'] = budget_data.get('budget_total', 0)
        ws[f'A{row}'].font = ExcelGenerator.TOTAL_FONT
        ws[f'E{row}'].font = ExcelGenerator.TOTAL_FONT
        ws[f'A{row}'].fill = ExcelGenerator.TOTAL_FILL
        ws[f'E{row}'].fill = ExcelGenerator.TOTAL_FILL
        ws[f'E{row}'].number_format = '¥#,##0'
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        row += 1
        
        # 粗利
        profit_rate = budget_data.get('profit_rate', 0)
        ws[f'A{row}'] = f"粗利率 {profit_rate}%"
        ws[f'E{row}'] = budget_data.get('profit_amount', 0)
        ws[f'A{row}'].font = Font(name='游ゴシック', size=10, bold=True)
        ws[f'E{row}'].font = Font(name='游ゴシック', size=10, bold=True)
        ws[f'E{row}'].number_format = '¥#,##0'
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        row += 1
        
        # 見積金額
        ws[f'A{row}'] = "見積金額"
        ws[f'E{row}'] = budget_data.get('estimate_amount', 0)
        ws[f'A{row}'].font = Font(name='游ゴシック', size=12, bold=True, color="0066cc")
        ws[f'E{row}'].font = Font(name='游ゴシック', size=12, bold=True, color="0066cc")
        ws[f'E{row}'].number_format = '¥#,##0'
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        
        # 保存
        wb.save(output_path)
        
        return output_path
    
    @staticmethod
    def generate_project_list_excel(projects: List[Dict[str, Any]], output_path: Path) -> Path:
        """
        工事台帳一覧をExcel出力
        
        Args:
            projects: プロジェクトリスト
            output_path: 出力先パス
        
        Returns:
            出力ファイルのパス
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工事台帳一覧"
        
        # 列幅設定
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        
        row = 1
        
        # タイトル
        ws.merge_cells(f'A{row}:H{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "工事台帳一覧"
        title_cell.font = Font(name='游ゴシック', size=18, bold=True, color="0066cc")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
        row += 2
        
        # ヘッダー
        headers = ['工事名', '発注者', '契約金額', '実行予算', '実績原価', '粗利率', '進捗', '状態']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row, col_idx)
            cell.value = header
            cell.font = ExcelGenerator.HEADER_FONT
            cell.fill = ExcelGenerator.HEADER_FILL
            cell.border = ExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # データ
        for project in projects:
            ws[f'A{row}'] = project.get('name', '')
            ws[f'B{row}'] = project.get('client', '')
            ws[f'C{row}'] = project.get('contract_amount', 0)
            ws[f'D{row}'] = project.get('budget_amount', 0)
            ws[f'E{row}'] = project.get('actual_cost', 0)
            ws[f'F{row}'] = project.get('profit_rate', 0)
            ws[f'G{row}'] = project.get('progress', 0)
            ws[f'H{row}'] = project.get('status', '')
            
            # 数値フォーマット
            ws[f'C{row}'].number_format = '¥#,##0'
            ws[f'D{row}'].number_format = '¥#,##0'
            ws[f'E{row}'].number_format = '¥#,##0'
            ws[f'F{row}'].number_format = '#0.0%'
            ws[f'G{row}'].number_format = '#0%'
            
            # 右寄せ
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
            
            ws[f'H{row}'].alignment = Alignment(horizontal='center')
            
            # ステータスの色付け
            status = project.get('status', '')
            if status == 'active':
                ws[f'H{row}'].fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                ws[f'H{row}'].font = Font(name='游ゴシック', size=10, bold=True, color="FFFFFF")
            elif status == 'warning':
                ws[f'H{row}'].fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
                ws[f'H{row}'].font = Font(name='游ゴシック', size=10, bold=True, color="FFFFFF")
            elif status == 'danger':
                ws[f'H{row}'].fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
                ws[f'H{row}'].font = Font(name='游ゴシック', size=10, bold=True, color="FFFFFF")
            
            # 罫線
            for col_idx in range(1, 9):
                ws.cell(row, col_idx).border = ExcelGenerator.THIN_BORDER
            
            row += 1
        
        # 保存
        wb.save(output_path)
        
        return output_path

def generate_budget_excel(budget_data: Dict[str, Any], output_path: str) -> str:
    """実行予算書Excel生成（エントリーポイント）"""
    return str(ExcelGenerator.generate_budget_excel(budget_data, Path(output_path)))

def generate_project_list_excel(projects: List[Dict[str, Any]], output_path: str) -> str:
    """工事台帳一覧Excel生成（エントリーポイント）"""
    return str(ExcelGenerator.generate_project_list_excel(projects, Path(output_path)))
