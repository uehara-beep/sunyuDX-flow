"""
KAKUSA形式Excel生成ユーティリティ
見積書・内訳書・実行予算内訳書・請求明細書を生成
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List
from app.models.estimate import Estimate, EstimateItem, EstimateDetail, CompanyMaster
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class KakusaExcelGenerator:
    """KAKUSA形式Excel生成クラス"""
    
    def __init__(self, company: CompanyMaster):
        self.company = company
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # デフォルトシートを削除
        
        # スタイル定義
        self.title_font = Font(name='MS Pゴシック', size=20, bold=True)
        self.header_font = Font(name='MS Pゴシック', size=11, bold=True)
        self.normal_font = Font(name='MS Pゴシック', size=10)
        self.company_font = Font(name='MS Pゴシック', size=9)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
    
    def generate(self, estimate: Estimate, output_path: str):
        """
        KAKUSA形式のExcelファイルを生成
        
        Args:
            estimate: 見積データ
            output_path: 出力先パス
        """
        logger.info(f"KAKUSA Excel生成開始: {estimate.estimate_number}")
        
        # 1. 御見積書
        self._create_estimate_sheet(estimate)
        
        # 2. 内訳明細書
        if estimate.details:
            self._create_detail_sheet(estimate)
        
        # 3. 実行予算内訳書
        self._create_budget_sheet(estimate)
        
        # 4. 請求明細書
        self._create_invoice_sheet(estimate)
        
        # 保存
        self.wb.save(output_path)
        logger.info(f"KAKUSA Excel生成完了: {output_path}")
    
    def _create_estimate_sheet(self, estimate: Estimate):
        """
        御見積書シートを作成（KAKUSA 1ページ目形式）
        """
        ws = self.wb.create_sheet("御見積書")
        
        # ===== ヘッダー部分 =====
        # タイトル
        ws['B2'] = '御　見　積　書'
        ws['B2'].font = self.title_font
        ws.merge_cells('B2:E2')
        
        # 見積番号・日付
        ws['H2'] = f"No. {estimate.estimate_number}"
        ws['H2'].font = self.normal_font
        ws['H3'] = estimate.estimate_date
        ws['H3'].font = self.normal_font
        
        # 顧客情報
        ws['B4'] = f"{estimate.customer_name}　御中"
        ws['B4'].font = self.header_font
        
        # 挨拶文
        ws['B6'] = '毎度、格別の御引立を賜り有難うございます。'
        ws['B7'] = '御依頼を戴きました本件に付き、誠心誠意検討を加え御見積りいたしましたので'
        ws['B8'] = '是非御下命戴けます様、お願い申し上げます。'
        
        # 工事情報
        row = 10
        ws[f'B{row}'] = '【工   事   名】'
        ws[f'B{row}'].font = self.header_font
        ws[f'F{row}'] = estimate.project_name
        row += 1
        
        ws[f'B{row}'] = '【工 事 場 所】'
        ws[f'B{row}'].font = self.header_font
        ws[f'F{row}'] = estimate.project_location or ''
        row += 1
        
        ws[f'B{row}'] = '【工　　期】'
        ws[f'B{row}'].font = self.header_font
        period = ''
        if estimate.project_period_from and estimate.project_period_to:
            period = f"自 {estimate.project_period_from} 至 {estimate.project_period_to}"
        ws[f'F{row}'] = period
        row += 1
        
        ws[f'B{row}'] = '【有 効 期 限】'
        ws[f'B{row}'].font = self.header_font
        ws[f'F{row}'] = estimate.valid_period
        row += 1
        
        ws[f'B{row}'] = '【支払い条件】'
        ws[f'B{row}'].font = self.header_font
        ws[f'F{row}'] = estimate.payment_terms
        row += 1
        
        if estimate.waste_notice:
            ws[f'B{row}'] = '【産業廃棄物】'
            ws[f'B{row}'].font = self.header_font
            ws[f'F{row}'] = estimate.waste_notice
            row += 1
        
        if estimate.special_notes:
            ws[f'B{row}'] = '【特記事項】'
            ws[f'B{row}'].font = self.header_font
            ws[f'F{row}'] = estimate.special_notes
            row += 1
        
        ws[f'B{row}'] = '【担 当 者】'
        ws[f'B{row}'].font = self.header_font
        ws[f'F{row}'] = estimate.staff_name
        row += 2
        
        # 金額サマリー
        ws[f'B{row}'] = '小　計　金　額'
        ws[f'B{row}'].font = self.header_font
        ws[f'F{row}'] = f'¥{estimate.subtotal:,.0f}'
        ws[f'F{row}'].font = self.header_font
        ws[f'F{row}'].alignment = self.right_align
        row += 1
        
        ws[f'B{row}'] = f'消 費 税({int(estimate.tax_rate * 100)}%)'
        ws[f'B{row}'].font = self.header_font
        ws[f'F{row}'] = f'¥{estimate.tax_amount:,.0f}'
        ws[f'F{row}'].font = self.header_font
        ws[f'F{row}'].alignment = self.right_align
        row += 1
        
        ws[f'B{row}'] = '合　計　金　額'
        ws[f'B{row}'].font = Font(name='MS Pゴシック', size=14, bold=True)
        ws[f'F{row}'] = f'¥{estimate.total_amount:,.0f}'
        ws[f'F{row}'].font = Font(name='MS Pゴシック', size=14, bold=True)
        ws[f'F{row}'].alignment = self.right_align
        row += 2
        
        # 明細テーブルヘッダー
        headers = ['No.', '名　称', '仕様・規格・寸法', '数量', '単位', '単　価', '金　額', '備　考']
        header_cols = ['B', 'C', 'D', 'F', 'G', 'H', 'I', 'J']
        
        for col, header in zip(header_cols, headers):
            cell = ws[f'{col}{row}']
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        row += 1
        detail_start_row = row
        
        # 明細データ
        for item in estimate.items:
            ws[f'B{row}'] = item.item_number
            ws[f'C{row}'] = item.item_name
            ws[f'D{row}'] = item.specification or ''
            ws[f'F{row}'] = item.quantity
            ws[f'G{row}'] = item.unit
            ws[f'H{row}'] = item.unit_price
            ws[f'I{row}'] = f'=F{row}*H{row}'  # 数式
            ws[f'J{row}'] = item.remarks or ''
            
            # スタイル適用
            for col in header_cols:
                ws[f'{col}{row}'].border = self.thin_border
                ws[f'{col}{row}'].font = self.normal_font
            
            ws[f'F{row}'].number_format = '#,##0.0'
            ws[f'H{row}'].number_format = '#,##0'
            ws[f'I{row}'].number_format = '#,##0'
            
            row += 1
        
        # 合計行
        ws[f'C{row}'] = '合　計　（税抜）'
        ws[f'C{row}'].font = self.header_font
        ws[f'I{row}'] = f'=SUM(I{detail_start_row}:I{row-1})'
        ws[f'I{row}'].font = self.header_font
        ws[f'I{row}'].number_format = '#,##0'
        ws[f'I{row}'].border = self.thin_border
        
        # 会社情報（右下）
        company_row = row + 3
        ws[f'H{company_row}'] = self.company.company_name
        ws[f'H{company_row}'].font = Font(name='MS Pゴシック', size=12, bold=True)
        company_row += 1
        
        ws[f'H{company_row}'] = f"〒{self.company.postal_code}"
        ws[f'H{company_row}'].font = self.company_font
        company_row += 1
        
        ws[f'H{company_row}'] = self.company.address
        ws[f'H{company_row}'].font = self.company_font
        company_row += 1
        
        ws[f'H{company_row}'] = f"TEL:{self.company.tel}  FAX:{self.company.fax}"
        ws[f'H{company_row}'].font = self.company_font
        
        # 列幅調整
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
    
    def _create_detail_sheet(self, estimate: Estimate):
        """
        内訳明細書シートを作成（KAKUSA 2ページ目形式）
        """
        ws = self.wb.create_sheet("内訳明細書")
        
        # タイトル
        ws['A1'] = '内　訳　明　細　書'
        ws['A1'].font = self.title_font
        
        # ヘッダー
        headers = ['名          　　称', '規 　　    格', '数　量', '単　位', '単  価', '金   額', '備   考']
        header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        
        row = 2
        for col, header in zip(header_cols, headers):
            cell = ws[f'{col}{row}']
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        row += 1
        current_category = ""
        category_start_row = row
        
        # 明細データ
        for detail in estimate.details:
            # カテゴリが変わったら小計を出力
            if current_category and current_category != detail.category:
                # 小計行
                ws[f'A{row}'] = '小計'
                ws[f'F{row}'] = f'=SUM(F{category_start_row}:F{row-1})'
                ws[f'F{row}'].font = self.header_font
                ws[f'F{row}'].number_format = '#,##0'
                row += 2
                category_start_row = row
            
            # カテゴリ名（新しいカテゴリの場合）
            if current_category != detail.category:
                ws[f'A{row}'] = detail.category
                ws[f'A{row}'].font = self.header_font
                row += 1
                category_start_row = row
                current_category = detail.category
            
            # 明細行
            ws[f'A{row}'] = detail.item_name
            ws[f'B{row}'] = detail.specification or ''
            ws[f'C{row}'] = detail.quantity
            ws[f'D{row}'] = detail.unit
            ws[f'E{row}'] = detail.unit_price
            ws[f'F{row}'] = f'=C{row}*E{row}'
            ws[f'G{row}'] = detail.remarks or ''
            
            # スタイル
            for col in header_cols:
                ws[f'{col}{row}'].border = self.thin_border
                ws[f'{col}{row}'].font = self.normal_font
            
            ws[f'C{row}'].number_format = '#,##0.0'
            ws[f'E{row}'].number_format = '#,##0'
            ws[f'F{row}'].number_format = '#,##0'
            
            row += 1
        
        # 最終小計
        if current_category:
            ws[f'A{row}'] = '小計'
            ws[f'F{row}'] = f'=SUM(F{category_start_row}:F{row-1})'
            ws[f'F{row}'].font = self.header_font
            ws[f'F{row}'].number_format = '#,##0'
            row += 2
        
        # 諸経費・法定福利費（内訳明細書から計算）
        # ※実際の計算ロジックは要件に応じて調整
        
        # 列幅調整
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25
    
    def _create_budget_sheet(self, estimate: Estimate):
        """
        実行予算内訳書シートを作成（KAKUSA 5ページ目形式）
        """
        ws = self.wb.create_sheet("実行予算内訳書")
        
        # タイトル
        ws['A1'] = '実行予算内訳書'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # 工事情報
        row = 3
        ws[f'A{row}'] = '工事番号'
        ws[f'B{row}'] = estimate.estimate_number
        ws[f'D{row}'] = '予算番号'
        ws[f'E{row}'] = estimate.estimate_number.replace('SYT', 'BDG')
        row += 1
        
        ws[f'A{row}'] = '工事名'
        ws[f'B{row}'] = estimate.project_name
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
        
        ws[f'A{row}'] = '発注者'
        ws[f'B{row}'] = estimate.customer_name
        row += 1
        
        ws[f'A{row}'] = '契約日'
        ws[f'B{row}'] = estimate.estimate_date
        ws[f'D{row}'] = '完工予定日'
        ws[f'E{row}'] = estimate.project_period_to or ''
        row += 2
        
        ws[f'A{row}'] = '初期契約金額'
        ws[f'B{row}'] = estimate.total_amount
        ws[f'B{row}'].number_format = '¥#,##0'
        row += 1
        
        ws[f'A{row}'] = '増減金額'
        ws[f'B{row}'] = 0
        ws[f'B{row}'].number_format = '¥#,##0'
        row += 1
        
        ws[f'A{row}'] = '契約金額計'
        ws[f'B{row}'] = f'=B{row-2}+B{row-1}'
        ws[f'B{row}'].number_format = '¥#,##0'
        ws[f'B{row}'].font = self.header_font
        row += 1
        
        ws[f'A{row}'] = '今回予算計'
        ws[f'B{row}'] = estimate.subtotal
        ws[f'B{row}'].number_format = '¥#,##0'
        row += 1
        
        ws[f'A{row}'] = '予定粗利'
        ws[f'B{row}'] = f'=B{row-4}-B{row-1}'
        ws[f'B{row}'].number_format = '¥#,##0'
        ws[f'B{row}'].font = Font(name='MS Pゴシック', size=10, bold=True, color='0000FF')
        row += 1
        
        ws[f'A{row}'] = '予定粗利率'
        ws[f'B{row}'] = f'=B{row-1}/B{row-5}'
        ws[f'B{row}'].number_format = '0.0%'
        ws[f'B{row}'].font = Font(name='MS Pゴシック', size=10, bold=True, color='0000FF')
        row += 2
        
        # 予算明細ヘッダー
        headers = ['名　称', '仕　様', '数 量', '単位', '単　価', '金　額', '業者名']
        header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        
        for col, header in zip(header_cols, headers):
            cell = ws[f'{col}{row}']
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        row += 1
        budget_start_row = row
        
        # 費用分類別に集計（簡易版）
        # ※実際は estimate.details をベースに労務費・材料費・外注費・経費に分類
        cost_categories = {
            '労務費': 0,
            '材料費': 0,
            '外注費': estimate.subtotal * 0.7,  # 仮計算
            '経費': estimate.subtotal * 0.3      # 仮計算
        }
        
        for category, amount in cost_categories.items():
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = self.header_font
            ws[f'C{row}'] = 1
            ws[f'D{row}'] = '式'
            ws[f'F{row}'] = amount
            ws[f'F{row}'].number_format = '#,##0'
            
            for col in header_cols:
                ws[f'{col}{row}'].border = self.thin_border
            
            row += 1
        
        # 合計
        ws[f'A{row}'] = '【　合　計　】'
        ws[f'A{row}'].font = self.header_font
        ws[f'F{row}'] = f'=SUM(F{budget_start_row}:F{row-1})'
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].font = self.header_font
        ws[f'F{row}'].border = self.thin_border
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
    
    def _create_invoice_sheet(self, estimate: Estimate):
        """
        請求明細書シートを作成（KAKUSA 7ページ目形式）
        """
        ws = self.wb.create_sheet("請求明細書")
        
        # タイトル
        ws['A1'] = '御　請　求　書'
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # 請求番号・日付
        invoice_number = estimate.estimate_number.replace('SYT', 'INV')
        ws['G1'] = f"No. {invoice_number}"
        ws['G1'].font = self.normal_font
        ws['G2'] = f"取引年月日 {estimate.estimate_date}"
        ws['G2'].font = self.normal_font
        
        row = 3
        
        # 顧客情報
        ws[f'A{row}'] = f"{estimate.customer_name}　御中"
        ws[f'A{row}'].font = self.header_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 2
        
        # 挨拶文
        ws[f'A{row}'] = '毎度有難うございます。'
        ws[f'A{row+1}'] = '下記の通り、御請求申し上げます。'
        row += 3
        
        # 工事情報
        ws[f'A{row}'] = '工 事 名'
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'] = estimate.project_name
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
        
        ws[f'A{row}'] = '工事場所'
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'] = estimate.project_location or ''
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
        
        ws[f'A{row}'] = '支払期限'
        ws[f'A{row}'].font = self.header_font
        ws[f'B{row}'] = ''  # 後で設定可能
        row += 2
        
        # 御請求金額
        ws[f'A{row}'] = '御請求金額'
        ws[f'A{row}'].font = Font(name='MS Pゴシック', size=14, bold=True)
        ws[f'D{row}'] = f'¥{estimate.total_amount:,.0f}'
        ws[f'D{row}'].font = Font(name='MS Pゴシック', size=14, bold=True)
        ws[f'D{row}'].alignment = self.right_align
        row += 2
        
        # 明細ヘッダー
        headers = ['名　　称', '仕　　様', '数 量', '単 位', '単　価', '金　額']
        header_cols = ['A', 'B', 'C', 'D', 'E', 'F']
        
        for col, header in zip(header_cols, headers):
            cell = ws[f'{col}{row}']
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        row += 1
        detail_start_row = row
        
        # 明細データ
        for item in estimate.items:
            ws[f'A{row}'] = item.item_name
            ws[f'B{row}'] = item.specification or ''
            ws[f'C{row}'] = item.quantity
            ws[f'D{row}'] = item.unit
            ws[f'E{row}'] = item.unit_price
            ws[f'F{row}'] = f'=C{row}*E{row}'
            
            for col in header_cols:
                ws[f'{col}{row}'].border = self.thin_border
                ws[f'{col}{row}'].font = self.normal_font
            
            ws[f'C{row}'].number_format = '#,##0.0'
            ws[f'E{row}'].number_format = '#,##0'
            ws[f'F{row}'].number_format = '#,##0'
            
            row += 1
        
        # 合計
        ws[f'A{row}'] = '【　合　計　】'
        ws[f'A{row}'].font = self.header_font
        ws[f'F{row}'] = f'=SUM(F{detail_start_row}:F{row-1})'
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].border = self.thin_border
        row += 1
        
        # 消費税
        ws[f'A{row}'] = f'【消費税 {int(estimate.tax_rate*100)}%】'
        ws[f'A{row}'].font = self.header_font
        ws[f'F{row}'] = f'=F{row-1}*{estimate.tax_rate}'
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].border = self.thin_border
        row += 1
        
        # 総合計
        ws[f'A{row}'] = '【　総合計　】'
        ws[f'A{row}'].font = self.header_font
        ws[f'F{row}'] = f'=F{row-2}+F{row-1}'
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].font = self.header_font
        ws[f'F{row}'].border = self.thin_border
        row += 3
        
        # 振込先情報
        ws[f'A{row}'] = '振込先'
        ws[f'A{row}'].font = self.header_font
        row += 1
        
        ws[f'A{row}'] = f"{self.company.bank_name} {self.company.bank_branch}"
        row += 1
        ws[f'A{row}'] = f"{self.company.bank_account_type} {self.company.bank_account_number}"
        row += 1
        ws[f'A{row}'] = f"名義：{self.company.bank_account_name}"
        row += 2
        
        ws[f'A{row}'] = '備考'
        ws[f'A{row}'].font = self.header_font
        
        # 会社情報（右下）
        company_row = row + 2
        ws[f'E{company_row}'] = self.company.company_name
        ws[f'E{company_row}'].font = Font(name='MS Pゴシック', size=12, bold=True)
        company_row += 1
        
        ws[f'E{company_row}'] = f"〒{self.company.postal_code}"
        ws[f'E{company_row}'].font = self.company_font
        company_row += 1
        
        ws[f'E{company_row}'] = self.company.address
        ws[f'E{company_row}'].font = self.company_font
        company_row += 1
        
        ws[f'E{company_row}'] = f"TEL:{self.company.tel}  FAX:{self.company.fax}"
        ws[f'E{company_row}'].font = self.company_font
        company_row += 1
        
        ws[f'E{company_row}'] = f"登録番号: {self.company.registration_number}"
        ws[f'E{company_row}'].font = self.company_font
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
