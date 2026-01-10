"""
Excelファイル解析ユーティリティ
サンユウテック形式の見積Excelを解析してデータを抽出
"""
import openpyxl
from typing import Dict, List, Optional
from app.models.estimate import ParsedEstimateData, EstimateItem, EstimateDetail
import logging

logger = logging.getLogger(__name__)


class ExcelParser:
    """Excel解析クラス"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        
    def parse(self) -> ParsedEstimateData:
        """
        Excelファイルを解析してデータを抽出
        
        Returns:
            ParsedEstimateData: 抽出されたデータ
        """
        logger.info(f"Excel解析開始: {self.file_path}")
        
        # 表紙シートから基本情報を取得
        cover_data = self._parse_cover_sheet()
        
        # 内訳明細書シートから詳細情報を取得
        detail_data = self._parse_detail_sheet()
        
        # データを統合
        parsed_data = ParsedEstimateData(
            customer_name=cover_data.get("customer_name", ""),
            project_name=cover_data.get("project_name", ""),
            project_location=cover_data.get("project_location"),
            staff_name=cover_data.get("staff_name"),
            valid_period=cover_data.get("valid_period"),
            payment_terms=cover_data.get("payment_terms"),
            waste_notice=cover_data.get("waste_notice"),
            special_notes=cover_data.get("special_notes"),
            subtotal=cover_data.get("subtotal", 0.0),
            tax_amount=cover_data.get("tax_amount", 0.0),
            total_amount=cover_data.get("total_amount", 0.0),
            items=cover_data.get("items", []),
            details=detail_data.get("details", [])
        )
        
        logger.info(f"Excel解析完了: {parsed_data.project_name}")
        return parsed_data
    
    def _parse_cover_sheet(self) -> Dict:
        """
        表紙シートから基本情報を抽出
        
        Returns:
            Dict: 基本情報
        """
        try:
            ws = self.wb['表紙']
        except KeyError:
            ws = self.wb.active
            logger.warning("'表紙'シートが見つかりません。アクティブシートを使用します。")
        
        data = {}
        
        # 基本情報の抽出（セル位置はサンユウテック形式に基づく）
        try:
            # 顧客名 (B4セル)
            customer = ws['B4'].value
            if customer:
                data['customer_name'] = str(customer).strip()
            
            # 工事名 (F7セル)
            project = ws['F7'].value
            if project:
                data['project_name'] = str(project).strip()
            
            # 工事場所 (F8セル)
            location = ws['F8'].value
            if location:
                data['project_location'] = str(location).strip()
            
            # 有効期限 (F10セル)
            valid = ws['F10'].value
            if valid:
                data['valid_period'] = str(valid).strip()
            
            # 支払条件 (F11セル)
            payment = ws['F11'].value
            if payment:
                data['payment_terms'] = str(payment).strip()
            
            # 産廃特記 (F12セル)
            waste = ws['F12'].value
            if waste:
                data['waste_notice'] = str(waste).strip()
            
            # 特記事項 (F13セル)
            notes = ws['F13'].value
            if notes:
                data['special_notes'] = str(notes).strip()
            
            # 担当者 (F14セル)
            staff = ws['F14'].value
            if staff:
                data['staff_name'] = str(staff).strip()
            
            # 金額情報の抽出
            # 小計・消費税・合計の行を探す
            for row in range(20, 40):
                cell_b = ws.cell(row, 2).value  # B列
                cell_f = ws.cell(row, 6).value  # F列
                
                if cell_b and '小　計' in str(cell_b):
                    if cell_f and isinstance(cell_f, (int, float)):
                        data['subtotal'] = float(cell_f)
                
                elif cell_b and '消 費 税' in str(cell_b):
                    if cell_f and isinstance(cell_f, (int, float)):
                        data['tax_amount'] = float(cell_f)
                
                elif cell_b and '合　計' in str(cell_b):
                    if cell_f and isinstance(cell_f, (int, float)):
                        data['total_amount'] = float(cell_f)
            
            # 明細の抽出（表紙の簡易明細）
            items = []
            item_start_row = 27  # 明細開始行（目安）
            
            for row in range(item_start_row, min(item_start_row + 15, ws.max_row + 1)):
                # No. | 名称 | 仕様 | 数量 | 単位 | 単価 | 金額
                no_val = ws.cell(row, 2).value  # B列: No
                name_val = ws.cell(row, 4).value  # D列: 名称
                
                if not name_val or str(name_val).strip() in ['', '合　計']:
                    continue
                
                # 数値取得（エラー処理付き）
                try:
                    quantity = float(ws.cell(row, 6).value or 0)
                except:
                    quantity = 0.0
                
                unit = str(ws.cell(row, 7).value or '式').strip()
                
                try:
                    unit_price = float(ws.cell(row, 8).value or 0)
                except:
                    unit_price = 0.0
                
                try:
                    amount = float(ws.cell(row, 9).value or 0)
                except:
                    amount = quantity * unit_price if quantity and unit_price else 0.0
                
                spec = ws.cell(row, 5).value
                remarks = ws.cell(row, 10).value
                
                item = EstimateItem(
                    item_number=len(items) + 1,
                    item_name=str(name_val).strip(),
                    specification=str(spec).strip() if spec else None,
                    quantity=quantity,
                    unit=unit,
                    unit_price=unit_price,
                    amount=amount,
                    remarks=str(remarks).strip() if remarks else None
                )
                items.append(item)
            
            data['items'] = items
            
        except Exception as e:
            logger.error(f"表紙シート解析エラー: {e}")
            raise
        
        return data
    
    def _parse_detail_sheet(self) -> Dict:
        """
        内訳明細書シートから詳細情報を抽出
        
        Returns:
            Dict: 詳細情報
        """
        data = {"details": []}
        
        try:
            ws = self.wb['内訳明細書']
        except KeyError:
            logger.warning("'内訳明細書'シートが見つかりません。")
            return data
        
        try:
            current_category = ""
            
            # ヘッダー行をスキップして明細行を読み取る
            for row in range(3, min(100, ws.max_row + 1)):
                # A列: 名称
                name_val = ws.cell(row, 1).value
                
                if not name_val:
                    continue
                
                name_str = str(name_val).strip()
                
                # カテゴリ行の判定（数量が無い行）
                quantity_val = ws.cell(row, 3).value
                if not quantity_val and name_str and '工事' in name_str:
                    current_category = name_str
                    continue
                
                # 小計・注記行はスキップ
                if any(keyword in name_str for keyword in ['小　　　　　計', '※', '内　訳']):
                    continue
                
                # 明細行の解析
                if quantity_val:
                    try:
                        quantity = float(quantity_val)
                    except:
                        continue
                    
                    spec = ws.cell(row, 2).value
                    unit = str(ws.cell(row, 4).value or '式').strip()
                    
                    try:
                        unit_price = float(ws.cell(row, 5).value or 0)
                    except:
                        unit_price = 0.0
                    
                    try:
                        amount = float(ws.cell(row, 6).value or 0)
                    except:
                        amount = quantity * unit_price
                    
                    remarks = ws.cell(row, 7).value
                    
                    detail = EstimateDetail(
                        category=current_category,
                        item_name=name_str,
                        specification=str(spec).strip() if spec else None,
                        quantity=quantity,
                        unit=unit,
                        unit_price=unit_price,
                        amount=amount,
                        remarks=str(remarks).strip() if remarks else None,
                        item_order=len(data['details']) + 1
                    )
                    data['details'].append(detail)
            
        except Exception as e:
            logger.error(f"内訳明細書シート解析エラー: {e}")
        
        return data
    
    def close(self):
        """ワークブックを閉じる"""
        if self.wb:
            self.wb.close()
